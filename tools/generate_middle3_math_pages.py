from __future__ import annotations

import json
import re
from collections import defaultdict
from pathlib import Path
from urllib.parse import quote

from openpyxl import load_workbook

try:
    from . import generate_middle2_math_pages as base
except ImportError:
    import generate_middle2_math_pages as base


SITE = Path(__file__).resolve().parents[1]
DOMAIN = base.DOMAIN
SITE_NAME = base.SITE_NAME
MANUSCRIPT = SITE.parent / "참고자료" / "원고모음(엑셀)" / "중3 수학학원 원고.xlsx"
PARENT = "과목별학원"
CATEGORY = "중3수학학원"
CATEGORY_LABEL = "중3 수학학원"
PUBLISH_DATE = "2026-08-07"


def canonical(*parts: str) -> str:
    return DOMAIN + "/" + "/".join(quote(part, safe="") for part in parts) + "/"


def esc(value: object) -> str:
    return base.esc(value)


def row_value(row: dict[str, str], key: str) -> str:
    return base.row_value(row, key)


def choose(local: str, label: str, values):
    return base.choose(local, f"{CATEGORY}:{label}", values)


def load_manuscripts() -> list[str]:
    """371개 중3 수학 원고를 센터 CSV와 같은 행 순서로 읽습니다."""
    if not MANUSCRIPT.exists():
        raise FileNotFoundError(MANUSCRIPT)
    workbook = load_workbook(MANUSCRIPT, read_only=True, data_only=True)
    sheet = workbook.active
    values = [str(row[0] or "").strip() for row in sheet.iter_rows(values_only=True)]
    workbook.close()
    return values


def sanitize_middle3_manuscript(value: str, row: dict[str, str]) -> str:
    """외부 원고의 개별 서술은 유지하며 광고성·근거 없는 표현을 중립화합니다."""
    value = base.sanitize_manuscript(value, row)

    # base 정제기가 교차과목 문장을 대체하며 삽입한 중2 전용 표현만 중3 맥락으로 바꿉니다.
    # Excel 원문의 학년 숫자를 일괄 치환하지 않으므로 중1·중2 누적 개념 언급은 그대로 남습니다.
    replacements = {
        "중2 수학은 현재 진도만 앞서가기보다": "중3 수학은 고교 선행 범위만 넓히기보다",
        "같은 중2 과정이라도": "같은 중3 과정이라도",
        "중2 수학의 개념·유형·서술형·오답": "중3 수학의 누적 개념·내신·서술형·오답",
        "중2 수학": "중3 수학",
    }
    for before, after in replacements.items():
        value = value.replace(before, after)

    quality_replacements = {
        "바로 점수": "즉시 결과",
        "점수 만드는": "풀이 과정을 점검하는",
        "정답률이 오르": "정확도가 달라지",
        "단기간 성적": "짧은 기간의 결과",
        "단기간 점수": "짧은 기간의 결과",
        "성적이 안 오르": "학습 변화가 더딘",
        "성적이 오르": "풀이가 안정되",
        "성적에 필요한": "학습에 필요한",
        "상승 로드맵": "점검 로드맵",
        "난도 상승": "난도 확장",
    }
    for before, after in quality_replacements.items():
        value = value.replace(before, after)

    local = row_value(row, "근처 수업가능 동네")
    neutral_results = [
        "현재 풀이 기록을 기준으로 다음 확인 범위를 조정합니다.",
        "누적 공백과 재풀이 결과를 나누어 다음 학습 순서를 정합니다.",
        "정확도와 풀이 시간을 따로 기록해 고교 진입 전 보완 순서를 살핍니다.",
        "특정 결과를 단정하지 않고 학생이 혼자 해결한 범위를 기준으로 계획을 조정합니다.",
    ]

    def replace_result(match: re.Match[str]) -> str:
        return choose(local, f"manuscript-result-{match.group(0)}", neutral_results)

    value = re.sub(
        r"[^<>.!?]*(?:(?:성적|점수|등급|합격)[^<>.!?]{0,35}(?:보장|상승|향상|오르|올리|끌어올리|완성|달성)|(?:보장|상승|향상|오르|올리|끌어올리)[^<>.!?]{0,25}(?:성적|점수|등급|합격))[^<>.!?]*(?:[.!?]|$)",
        replace_result,
        value,
        flags=re.I,
    )
    school_neutral = [
        "학교별 경향을 임의로 단정하지 않고 학생이 가져온 시험 범위와 학교 자료를 확인합니다.",
        "현재 학교의 교과서·프린트·최근 시험지에서 확인되는 범위만 학습 계획에 반영합니다.",
        "학교 자료가 없는 내용은 추정하지 않고 상담에서 실제 진도와 시험 범위를 먼저 확인합니다.",
    ]

    def replace_school(match: re.Match[str]) -> str:
        return choose(local, f"manuscript-school-{match.group(0)}", school_neutral)

    value = re.sub(
        r"[^<>.!?]*(?:(?:학교별|지역\s*학생들의?)[^<>.!?]{0,35}(?:출제\s*(?:흐름|경향)|시험\s*범위))[^<>.!?]*(?:맞춰|반영|분석|대비)[^<>.!?]*(?:[.!?]|$)",
        replace_school,
        value,
        flags=re.I,
    )
    return value


def build_description(row: dict[str, str], profile: dict[str, str], page_index: int) -> str:
    """지역 사실을 제외해도 64가지 조합으로 구분되는 설명을 만듭니다."""
    local = row_value(row, "근처 수업가능 동네")
    region = row_value(row, "지역")
    district = row_value(row, "시or구")
    center = row_value(row, "센터명") or f"{local} 학습센터"
    openings = [
        "현재 학교 진도와 누적 개념 공백을 함께 확인하고",
        "최근 답안의 조건 해석·계산·서술형 오류를 나누어 살피며",
        "중3 내신 준비와 고교 진입 전 개념 연결을 구분해 점검하고",
        "독립 풀이·간격 재풀이·검산 기록을 기준으로 비교하며",
        "시험 범위 확정 전후의 복습 계획과 완료 기록을 확인하고",
        "현재 단원을 막는 중1·중2 개념과 시간 관리를 함께 살피며",
        "학교 자료의 풀이 흔적과 해설 없는 재풀이 결과를 대조하고",
        "서술형 근거·오답 원인·주간 복습 가능 시간을 구체화하며",
    ]
    closings = [
        "상담 전에 준비할 자료와 운영 확인 항목을 정리했습니다.",
        "학생 상황에 맞는 시작점과 다음 확인 순서를 안내합니다.",
        "과장된 결과 대신 실제 기록으로 비교할 기준을 담았습니다.",
        "수업 가능 학년과 공개 센터 정보를 함께 확인할 수 있습니다.",
        "선행 범위보다 우선 보완할 학습 단계를 판단하도록 구성했습니다.",
        "평상시 학습과 시험기 계획을 나누어 질문할 수 있게 정리했습니다.",
        "재풀이 결과가 다음 계획에 반영되는지 살펴볼 기준을 제시합니다.",
        "학교 정보는 공개된 자료만 사용하고 미기재 항목은 구분했습니다.",
    ]
    opening = openings[page_index % len(openings)]
    closing = closings[(page_index // len(openings)) % len(closings)]
    return f"{region} {district} {local} 중3 수학학원 선택을 위해 {center} 공개 정보를 바탕으로 {opening} {closing}"


def diversify_paragraphs(markup: str, row: dict[str, str], page_index: int, offset: int = 0) -> str:
    """공통 학습 원칙을 페이지별 실제 확인 문맥과 결합합니다."""
    signals = [
        "처음 풀이를 멈춘 지점", "답을 가린 뒤 다시 푼 결과", "조건을 식으로 옮긴 과정",
        "부호와 계산 순서를 검산한 흔적", "서술형에서 빠뜨린 근거", "현재 단원을 막는 이전 개념",
        "해설을 보기 전 남긴 시도", "문제별 풀이 시간과 검산 시간", "질문 표시가 남은 교재 부분",
        "기본 유형과 변형 문제의 차이", "학교 진도와 누적 복습의 간격", "주간 계획의 실제 완료 여부",
    ]
    actions = [
        "별도 표시한 뒤", "최근 답안과 나란히 둔 뒤", "오류 원인별로 분류한 뒤",
        "학생의 설명과 대조한 뒤", "완료 기준에 따라 구분한 뒤", "다음 질문 항목으로 옮긴 뒤",
        "독립 풀이 여부와 비교한 뒤", "주간 기록에 연결한 뒤", "재풀이 날짜와 함께 적은 뒤",
        "학교 자료에서 다시 찾은 뒤", "비슷한 새 문제에 적용한 뒤", "우선순위 표에 반영한 뒤",
    ]
    timings = [
        "수업 직후와 주중의 차이를 확인합니다.", "다음 재풀이 날짜에 같은 기준으로 확인합니다.",
        "한 주 뒤 해설 없는 풀이와 다시 비교합니다.", "시험 범위 확정 전후의 변화를 나누어 봅니다.",
        "다음 계획에서 유지할 항목과 바꿀 항목을 정합니다.", "학생이 혼자 설명할 수 있는 범위를 다시 확인합니다.",
        "학교 진도와 누적 복습의 분량을 따로 조정합니다.", "정확도가 안정된 뒤 시간 기준을 추가합니다.",
        "상담에서 실제 관리 기록의 반영 방식을 질문합니다.", "현재 단원에 바로 적용되는지 결과를 남깁니다.",
        "서술형 답안과 일반 풀이에서 각각 재확인합니다.", "완료되지 않은 항목만 다음 주 계획으로 넘깁니다.",
    ]
    paragraph_index = 0

    def add_context(match: re.Match[str]) -> str:
        nonlocal paragraph_index
        attrs, body = match.group(1) or "", match.group(2)
        plain = base.strip_tags(body)
        current = paragraph_index
        paragraph_index += 1
        if len(plain) < 45:
            return match.group(0)
        code = (page_index * 37 + current + offset) % (len(signals) * len(actions) * len(timings))
        signal = signals[code % len(signals)]
        action = actions[(code // len(signals)) % len(actions)]
        timing = timings[(code // (len(signals) * len(actions))) % len(timings)]
        context = f"확인 기록에는 {base.object_form(signal)} {action} {timing}"
        return f"<p{attrs}>{body.rstrip()} {esc(context)}</p>"

    return re.sub(r"<p(\s[^>]*)?>(.*?)</p>", add_context, markup, flags=re.I | re.S)


def page_profile(row: dict[str, str]) -> dict[str, str]:
    local = row_value(row, "근처 수업가능 동네")
    students = [
        "중1·중2 개념은 배웠지만 현재 문제에서 어떤 개념을 꺼내야 하는지 판단이 느린 학생",
        "풀이 과정은 이해하지만 며칠 뒤 혼자 풀 때 첫 단계가 이어지지 않는 학생",
        "계산은 빠르지만 조건 표시와 검산을 생략해 작은 오류가 반복되는 학생",
        "시험 범위가 정해진 뒤에도 단원별 복습 순서를 잡지 못하는 학생",
        "서술형에서 식은 세우지만 사용한 개념과 근거 문장을 충분히 남기지 못하는 학생",
        "현재 진도와 이전 학년의 누적 공백을 함께 구분해야 하는 학생",
        "문제량은 충분하지만 오답을 답만 고치고 재풀이 날짜를 정하지 않는 학생",
        "고교 수학을 서두르기보다 중학교 과정의 연결 상태를 먼저 확인해야 하는 학생",
        "시간 제한이 생기면 검산 순서가 흔들리고 쉬운 문제에서도 실수가 늘어나는 학생",
        "기본 유형은 풀지만 조건이 바뀐 문제에서 풀이 전략을 다시 세우기 어려운 학생",
        "학교 진도를 따라가면서 누적 복습 시간을 따로 확보하기 어려운 학생",
        "계획은 세우지만 완료 여부와 질문할 문제를 기록하지 않아 복습이 밀리는 학생",
        "설명을 들으면 이해하지만 풀이를 말과 식으로 다시 표현하는 연습이 부족한 학생",
        "시험 직전에만 오답을 모아 평상시 학습 기록이 남지 않는 학생",
    ]
    priorities = [
        "현재 답안에서 개념·조건 해석·계산·검산 오류를 먼저 구분합니다.",
        "중1·중2 누적 개념 중 현재 진도를 막는 연결 지점부터 확인합니다.",
        "정답보다 풀이의 첫 단계와 사용한 근거를 학생이 설명할 수 있는지 봅니다.",
        "기본 유형의 간격 재풀이가 안정된 뒤 조건이 달라진 문제로 확장합니다.",
        "학교 진도와 누적 오답을 한 표에 놓고 주간 완료 순서를 정합니다.",
        "시험 범위 확정 전과 확정 후의 복습 계획을 서로 다르게 구성합니다.",
        "서술형 답안에서 조건·식·근거·결론이 이어지는지 단계별로 점검합니다.",
        "정확한 풀이 순서가 자리 잡은 뒤 시간 제한과 검산 기준을 적용합니다.",
        "고교 진입 전 선행 범위보다 중학교 개념을 새 문제에 적용하는 힘을 봅니다.",
        "수업 직후·주중·시험 전으로 재풀이 간격과 완료 기록을 나눕니다.",
    ]
    checks = [
        "최근 시험지의 오답 원인을 학생이 말로 설명할 수 있는지",
        "현재 교재의 풀이 흔적에서 조건을 식으로 옮긴 과정이 남아 있는지",
        "해설 없이 다시 풀었을 때 같은 풀이 순서를 재현할 수 있는지",
        "중1·중2 개념 중 현재 단원을 막는 공백이 무엇인지",
        "서술형에서 사용한 성질과 결론 사이의 근거가 적혀 있는지",
        "주중에 학교 진도와 누적 복습을 각각 진행할 시간이 있는지",
        "문제별 풀이 시간과 검산 시간을 따로 기록하고 있는지",
        "질문할 문제와 혼자 해결한 문제를 구분해 표시하는지",
        "시험 범위가 확정된 뒤 단원별 완료 기준이 계획표에 있는지",
        "고교 과정에 앞서 중학교 수학의 핵심 연결을 스스로 설명할 수 있는지",
    ]
    bridges = [
        "중학교 개념을 외운 공식이 아니라 새로운 조건에 적용할 수 있는지 확인합니다.",
        "식의 변형과 함수 해석에서 조건, 식과 그래프 사이의 연결을 설명하게 합니다.",
        "도형 문제는 주어진 조건을 표시하고 사용할 성질과 결론을 순서대로 적게 합니다.",
        "자료와 확률 문제는 경우를 나누는 기준과 수치의 의미를 함께 확인합니다.",
        "고교 선행 전에 중학교 과정의 계산 정확도와 개념 연결 상태를 먼저 정리합니다.",
        "현재 단원을 막는 이전 개념만 골라 복원한 뒤 바로 새 문제에 적용합니다.",
        "정답을 가린 재풀이와 조건을 바꾼 문제를 구분해 독립 해결 범위를 봅니다.",
        "풀이를 말로 설명한 내용과 종이에 쓴 식이 같은 흐름인지 대조합니다.",
        "서술형은 계산 결과뿐 아니라 주어진 조건과 사용한 근거가 드러나게 구성합니다.",
        "평상시 누적 복습과 시험 범위 복습을 별도 일정으로 관리합니다.",
    ]
    rhythms = [
        "수업 직후 풀이 정리 → 주중 재풀이 → 한 주 뒤 누적 확인",
        "학교 진도 확인 → 우선 개념 복원 → 적용 문제 → 서술형 재확인",
        "오답 원인 분류 → 기본 유형 재현 → 조건 변형 → 검산 기록",
        "현재 범위 복습 → 이전 공백 연결 → 독립 풀이 → 계획 갱신",
        "정확도 확인 → 풀이 시간 기록 → 검산 시간 확보 → 누적 점검",
        "시험지 분석 → 단원별 완료 기준 → 학교 자료 적용 → 오답 재확인",
        "질문 표시 → 개념 설명 → 답을 가린 재풀이 → 새 문항 적용",
        "중학교 과정 연결 → 학교 진도 유지 → 서술형 근거 → 고교 준비 점검",
        "교재 진도 기록 → 막힌 단계 표시 → 주간 복습 → 완료 여부 확인",
        "평상시 공백 보완 → 범위 확정 후 집중 복습 → 시험 뒤 오류 갱신",
    ]
    return {
        "student": choose(local, "student", students),
        "priority": choose(local, "priority", priorities),
        "check": choose(local, "check", checks),
        "bridge": choose(local, "bridge", bridges),
        "rhythm": choose(local, "rhythm", rhythms),
    }


def contextualize_manuscript(raw: str, row: dict[str, str], repeated: set[str]) -> str:
    """반복 서명이 확인된 문단에만 검증 가능한 지역·학생 문맥을 더합니다."""
    local = row_value(row, "근처 수업가능 동네")
    region = row_value(row, "지역")
    district = row_value(row, "시or구")
    center = row_value(row, "센터명") or f"{local} 학습센터"
    schools = base.split_items(row_value(row, "타깃학교\n(중)"))
    evidence = schools[base.seed_for(CATEGORY, local, "manuscript-school") % len(schools)] if schools else "현재 학교 자료"
    profile = page_profile(row)

    raw = sanitize_middle3_manuscript(raw.strip(), row)
    raw = re.sub(
        r'^\s*<main class="article-main">',
        '<section class="shell academy-section article-main manuscript-panel">',
        raw,
        count=1,
        flags=re.I,
    )
    raw = re.sub(r"</main>\s*$", "</section>", raw, count=1, flags=re.I)
    heading = f"{region} {district} {local} 중3 수학 학습 설계"
    raw = re.sub(r"<h1>.*?</h1>", f"<h2>{esc(heading)}</h2>", raw, count=1, flags=re.I | re.S)

    templates = [
        "{local}에서는 {evidence}의 최근 풀이 흔적과 중1·중2 누적 공백을 함께 놓고 중3 복습 순서를 정하는 것이 좋습니다.",
        "이 항목은 {local} 학생의 {student} 상황과 연결해 확인하면 우선순위가 더 분명해집니다.",
        "{center} 상담에서는 {evidence} 자료로 독립 풀이와 설명을 들은 뒤의 풀이를 구분해 볼 수 있습니다.",
        "{region} {district} {local}에서는 고교 선행 범위보다 {evidence}에서 확인되는 현재 진도와 재풀이 기록을 먼저 비교해야 합니다.",
        "{local} 학부모는 이 기준을 볼 때 {check}를 질문으로 준비할 수 있습니다.",
        "같은 중3 과정이라도 {local} 학생의 학교 진도와 주간 복습 가능 시간에 따라 먼저 보완할 내용은 달라질 수 있습니다.",
        "{evidence}의 오답을 개념·조건 해석·계산·시간 배분으로 나누면 고교 진입 전에 남은 공백을 구체적으로 확인할 수 있습니다.",
        "{district} 지역 상담에서는 시험 범위 확정 전 누적 복습과 확정 후 내신 계획을 분리해 설명하는지 살펴볼 수 있습니다.",
        "{local}의 실제 주간표에는 {evidence} 점검 결과와 다음 재풀이 날짜가 함께 기록되어야 합니다.",
        "서술형은 답만 맞히는지보다 {local} 학생이 사용한 개념과 조건을 스스로 설명하는지까지 확인해야 합니다.",
    ]
    paragraph_index = 0

    def add_context(match: re.Match[str]) -> str:
        nonlocal paragraph_index
        attrs, body = match.group(1) or "", match.group(2)
        signature = base.paragraph_signature(body, local)
        current_index = paragraph_index
        paragraph_index += 1
        if signature not in repeated or len(signature) < 35:
            return match.group(0)
        template = templates[base.seed_for(CATEGORY, local, str(current_index), signature) % len(templates)]
        context = template.format(
            local=local,
            region=region,
            district=district,
            center=center,
            evidence=evidence,
            student=profile["student"],
            check=profile["check"],
        )
        return f"<p{attrs}>{body.rstrip()} {esc(context)}</p>"

    return re.sub(r"<p(\s[^>]*)?>(.*?)</p>", add_context, raw, flags=re.I | re.S)


def build_learning_article(row: dict[str, str], profile: dict[str, str], page_index: int) -> str:
    local = row_value(row, "근처 수업가능 동네")
    region = row_value(row, "지역")
    district = row_value(row, "시or구")
    center = row_value(row, "센터명") or f"{local} 학습센터"
    schools = base.split_items(row_value(row, "타깃학교\n(중)"))
    evidence = schools[base.seed_for(CATEGORY, local, "evidence") % len(schools)] if schools else "현재 학교의 시험지와 교재"
    openings = [
        f"{local}에서는 선행 분량보다 최근 답안에서 풀이가 멈춘 지점을 먼저 확인합니다.",
        f"{region} {district} {local}의 중3 과정은 학교 진도와 누적 복습을 서로 다른 일정으로 관리할 필요가 있습니다.",
        f"{center} 상담에서는 설명을 들은 문제와 해설 없이 해결한 문제를 구분해 시작점을 정합니다.",
        f"{evidence}를 기준으로 현재 단원과 이전 개념, 서술형 답안의 연결 상태를 함께 살펴봅니다.",
        f"중3 수학은 내신과 고교 준비가 겹치므로 {local} 학생의 실제 주간 복습 시간을 먼저 계산합니다.",
        f"{local} 학습 기록에서는 문제 수보다 오류 원인과 재풀이 결과가 다음 계획으로 이어지는지 봅니다.",
        f"시험 범위가 정해지기 전에는 {evidence}의 진도와 누적 공백을 분리해 적는 것이 좋습니다.",
        f"{district}에서 중3 수학을 비교할 때는 정답률보다 풀이 근거와 검산 순서가 재현되는지 확인합니다.",
        f"{center}의 공개 정보와 학생의 실제 교재 기록을 함께 놓아야 상담 질문을 구체적으로 정할 수 있습니다.",
        f"{local} 중3 학생은 평상시 복습과 시험기 복습의 완료 기준을 다르게 세울 필요가 있습니다.",
        f"{evidence}의 맞힌 문제도 근거를 설명하지 못한다면 독립 풀이 범위에서 따로 구분합니다.",
        f"고교 과정에 들어가기 전 {local}에서는 중학교 개념을 새 조건에 적용할 수 있는지 먼저 살펴봅니다.",
    ]
    diagnostic_domains = [
        ("누적 개념 연결", "현재 문제를 막는 이전 개념만 골라 복원하고 같은 날 현행 문제에 적용합니다."),
        ("조건 해석", "문제에서 주어진 조건과 구해야 할 값을 표시하고 사용할 개념을 고릅니다."),
        ("식 세우기", "계산을 시작하기 전에 조건이 어떤 식으로 이어지는지 학생이 말하게 합니다."),
        ("부호와 계산", "부호·괄호·계산 순서를 나누어 반복 오류가 시작된 단계를 찾습니다."),
        ("함수 표현", "표·식·그래프가 나타내는 같은 관계를 서로 바꾸어 설명하게 합니다."),
        ("그래프 해석", "축과 변화량, 교점의 의미를 문제 조건과 연결해 읽는지 확인합니다."),
        ("도형 조건", "주어진 길이·각·평행 관계를 그림에 직접 표시한 뒤 성질을 선택합니다."),
        ("도형 추론", "사용한 성질과 결론 사이에 빠진 근거가 없는지 순서대로 확인합니다."),
        ("확률과 경우", "경우를 나누는 기준이 겹치거나 빠지지 않았는지 표와 식으로 대조합니다."),
        ("자료 해석", "수치의 크기만 보지 않고 대표값과 분포가 의미하는 차이를 설명하게 합니다."),
        ("서술형 근거", "조건·개념·식·결론이 답안에서 자연스럽게 이어지는지 점검합니다."),
        ("풀이 시간", "정확한 풀이가 자리 잡은 뒤 문제 시간과 검산 시간을 따로 기록합니다."),
        ("검산 순서", "답의 범위와 조건 충족 여부를 일정한 순서로 다시 확인합니다."),
        ("오답 분류", "개념·조건·계산·시간 중 어느 원인인지 나누어 다음 행동을 정합니다."),
        ("질문 기록", "해설을 보기 전 시도와 막힌 부분을 표시해 필요한 설명 범위를 좁힙니다."),
        ("독립 재풀이", "수업 뒤 간격을 두고 같은 풀이를 해설 없이 재현할 수 있는지 봅니다."),
    ]
    actions = [
        "최근 시험지와 현재 교재를 나란히 놓고 차이를 기록합니다.", "맞힌 문제와 틀린 문제의 첫 풀이 단계를 비교합니다.",
        "학생의 말 설명과 종이에 남은 식이 같은 흐름인지 대조합니다.", "답을 가린 뒤 새 종이에 풀이 순서를 다시 구성합니다.",
        "숫자와 조건을 바꾼 문제에서도 같은 개념을 선택하는지 봅니다.", "오답 원인을 한 단어가 아니라 다음 행동까지 적게 합니다.",
        "학교 진도와 누적 공백을 한 표에 두되 완료 날짜는 나눕니다.", "설명을 들은 문제와 혼자 해결한 문제를 서로 다른 표시로 남깁니다.",
        "질문할 부분을 먼저 표시한 뒤 필요한 개념만 짧게 복원합니다.", "기본 유형이 안정된 뒤 조건이 달라진 문제로 범위를 넓힙니다.",
        "수업 직후와 며칠 뒤 결과를 비교해 기억과 이해를 구분합니다.", "풀이 시간보다 정확한 순서가 먼저 자리 잡았는지 확인합니다.",
        "서술형 답안을 조건 누락·개념·계산·결론으로 나눠 고칩니다.", "완료된 문제와 다시 볼 문제를 주간 계획표에 각각 반영합니다.",
        "학교 자료에서 확인되는 범위만 사용하고 미확인 내용은 추정하지 않습니다.", "다음 수업에서 재풀이 결과가 실제 과제 조정에 반영됐는지 봅니다.",
    ]
    verifications = [
        "다음 재풀이 날짜에 같은 오류가 남는지 확인합니다.", "한 주 뒤 학생이 근거를 다시 설명하는지 살펴봅니다.",
        "현재 단원의 새 문제에 바로 적용되는지 결과를 남깁니다.", "학교 범위 확정 전후의 완료 항목을 따로 비교합니다.",
        "주중 일정에서 실행할 수 있는 분량인지 다시 계산합니다.", "서술형과 일반 풀이에서 같은 개념을 사용하는지 대조합니다.",
        "검산 시간을 포함해 시험 안에서 가능한 순서인지 확인합니다.", "해설 없이 해결한 범위만 다음 단계의 시작점으로 삼습니다.",
        "틀린 이유와 고친 이유를 학생이 구분해 말하는지 봅니다.", "누적 복습이 학교 진도를 밀어내지 않는지 주간표로 확인합니다.",
        "상담에서 실제 관리 기록의 예시를 확인할 질문으로 바꿉니다.", "다음 과제에 같은 기준이 이어지는지 완료 기록을 확인합니다.",
        "고교 선행보다 중학교 개념 연결이 먼저 안정됐는지 봅니다.", "문제 수가 아니라 독립 해결 범위가 넓어졌는지 비교합니다.",
        "미완료 항목만 다음 주 계획으로 넘겨 과도한 반복을 줄입니다.", "학생이 사용할 검산 기준을 직접 말할 수 있는지 확인합니다.",
    ]

    def combined_cards(bank: list[tuple[str, str]], count: int, code_offset: int, css: str) -> str:
        cards: list[str] = []
        for slot in range(count):
            code = base.seed_for(CATEGORY, local, "combined-card", str(code_offset), str(slot))
            name, body = bank[code % len(bank)]
            action = actions[(code // len(bank)) % len(actions)]
            verify = verifications[(code // (len(bank) * len(actions))) % len(verifications)]
            cards.append(f'<article class="{css}"><h3>{esc(name)}</h3><p>{esc(body)} {esc(action)} {esc(verify)}</p></article>')
        return "".join(cards)

    first_cards = combined_cards(diagnostic_domains, 4, 0, "article-card")
    second_cards = combined_cards(diagnostic_domains, 4, 617, "article-card")
    plan_cards = combined_cards(diagnostic_domains, 4, 1231, "article-target-card")
    weekly_cards = combined_cards(diagnostic_domains, 5, 2099, "article-target-card")
    section_titles = [
        "최근 답안에서 중3 오류를 구분하는 기준", "현행 진도와 누적 공백을 함께 보는 방법",
        "서술형·시간 관리까지 연결하는 진단", "문제 수보다 풀이 기록을 먼저 보는 이유",
        "중3 내신 준비의 시작점을 좁히는 과정", "고교 진입 전에 확인할 수학 학습 신호",
        "학교 자료와 독립 풀이를 대조하는 기준", "오답 원인에 따라 복습 순서를 나누는 방법",
    ]
    transition_titles = [
        "평상시 복습과 시험기 계획을 나누는 흐름", "누적 개념을 현재 단원에 적용하는 순서",
        "내신·서술형·검산을 한 계획으로 잇는 방법", "중학교 개념에서 고교 준비로 넘어가는 기준",
        "재풀이 결과를 다음 주 계획에 반영하는 과정", "학교 진도와 시간 관리를 동시에 유지하는 방법",
        "설명 뒤 독립 풀이를 확인하는 네 단계", "시험 범위 확정 전후의 우선순위 조정",
    ]
    weekly_titles = [
        "이번 주 수학 기록에 남길 다섯 가지 기준", "진단 결과를 실제 복습으로 옮기는 주간 기준",
        "학생 상황에 맞춰 다음 확인일을 정하는 방법", "수업 뒤 독립 풀이를 확인하는 주간 기록",
        "시험 전까지 미완료 항목을 줄이는 실행 기준", "현재 답안에서 다음 과제를 정하는 확인 순서",
        "오답 원인별로 재풀이 날짜를 배치하는 방법", "학교 진도와 누적 복습을 함께 유지하는 기록",
    ]
    opening = choose(local, "article-opening", openings)
    section_title = choose(local, "article-section-title", section_titles)
    transition_title = choose(local, "article-transition-title", transition_titles)
    weekly_title = choose(local, "article-weekly-title", weekly_titles)
    evidence_intros = [
        f"{evidence}의 풀이 흔적에서 막힌 단계를 먼저 찾습니다.",
        f"{evidence}와 현재 교재를 함께 보며 완료 범위를 구분합니다.",
        f"{evidence}의 맞힌 문제도 근거를 설명할 수 있는지 확인합니다.",
        f"{evidence}에서 확인된 범위와 누적 오답을 별도로 기록합니다.",
        f"{evidence}를 준비해 설명 뒤 재풀이 결과까지 비교합니다.",
        f"{evidence}의 서술형과 일반 문항에서 반복 오류를 나눕니다.",
        f"{evidence} 기준으로 시험 전 완료할 항목을 다시 정합니다.",
        f"{evidence}와 주간표를 대조해 실행 가능한 복습량을 계산합니다.",
    ]
    evidence_intro = choose(local, "article-evidence-intro", evidence_intros)
    weekly_intros = [
        "최근 답안에서 확인된 문제만 골라 행동과 다음 확인일을 함께 적습니다.",
        "학생이 혼자 해결한 범위와 설명이 더 필요한 범위를 서로 다른 기록으로 남깁니다.",
        "학교 진도를 유지하면서도 이번 주 안에 다시 볼 누적 오답을 별도로 정합니다.",
        "정확도·풀이 근거·검산 순서를 각각 확인해 다음 과제의 분량을 조정합니다.",
        "시험 범위가 확정되기 전과 이후에 사용할 복습표를 나누어 준비합니다.",
        "고교 선행 여부보다 중학교 개념을 새 조건에 적용한 결과를 먼저 비교합니다.",
        "재풀이에서 다시 막힌 단계만 추려 다음 수업의 질문 순서를 만듭니다.",
        "교재의 완료 표시와 실제 해설 없는 풀이 결과가 일치하는지 대조합니다.",
    ]
    weekly_intro = choose(local, "article-weekly-intro", weekly_intros)
    transition_bases = [
        "최근 답안의 오류 단계와 실제 주간 시간을 함께 놓고 내신·누적 복습·고교 준비의 순서를 나눕니다.",
        "학교 진도에 필요한 복습과 고교 진입 전에 복원할 개념을 구분한 뒤 주간표에 배치합니다.",
        "독립 풀이 결과와 남은 학습 시간을 대조해 현재 단원, 누적 오답과 전환 준비의 우선순위를 정합니다.",
        "시험 범위가 정해지기 전의 공백 보완과 범위 확정 뒤의 내신 계획을 서로 다른 일정으로 관리합니다.",
        "풀이 정확도·서술 근거·소요 시간을 따로 기록해 이번 주에 먼저 끝낼 수학 과제를 좁힙니다.",
        "현재 교재의 진도와 다시 막힌 오답을 비교하고 고교 선행보다 먼저 확인할 중학교 개념을 고릅니다.",
        "학생이 설명한 풀이와 답을 가린 재풀이를 대조해 학교 학습과 누적 점검의 분량을 조정합니다.",
        "내신 대비 문제와 이전 단원 오답을 한꺼번에 늘리지 않고 완료 기준과 재확인 날짜를 각각 정합니다.",
        "조건 해석·계산·검산·시간 배분 중 반복되는 지점을 찾아 다음 수업과 주중 복습의 순서를 세웁니다.",
        "학교 자료에서 확인되는 범위와 학생의 실제 풀이 기록만 근거로 현재 학습과 전환 준비를 연결합니다.",
        "정답 여부보다 풀이의 첫 단계와 근거가 재현되는지 확인한 뒤 내신과 누적 복습의 비중을 정합니다.",
        "주간에 확보할 수 있는 시간 안에서 학교 진도, 미완료 오답과 고교 진입 전 점검을 순서대로 배치합니다.",
    ]
    transition_basis = choose(local, "article-transition-basis", transition_bases)
    closing_bank = [
        "진도표보다 독립 풀이 범위와 다음 재확인 날짜가 연결되는지 확인하는 편이 좋습니다.",
        "오답 원인과 고친 이유가 다음 과제에 반영되는지를 상담 기준으로 삼을 수 있습니다.",
        "학교 진도·누적 복습·고교 준비를 한꺼번에 늘리지 않고 완료 순서를 나누어야 합니다.",
        "실제 운영 일정과 학생의 주중 시간을 함께 놓아 실행 가능한 계획인지 살펴봐야 합니다.",
        "문제량보다 풀이 근거와 검산 순서가 며칠 뒤에도 재현되는지 확인할 필요가 있습니다.",
        "시험 범위가 확정되면 새 문제 추가보다 미완료 오답과 서술형 근거를 먼저 점검합니다.",
        "고교 선행 여부는 중학교 개념을 낯선 조건에 적용한 결과를 확인한 뒤 판단하는 것이 안전합니다.",
        "상담에서는 수업 설명보다 학생 기록이 다음 계획에 어떻게 반영되는지 구체적으로 질문해 보세요.",
    ]
    closing = choose(local, "article-closing", closing_bank)
    return f'''<section class="shell academy-section article-main manuscript-panel">
      <section class="article-hero"><p class="article-eyebrow">MIDDLE 3 MATH ROADMAP</p><h2>{esc(region)} {esc(district)} {esc(local)} 중3 수학 학습 설계</h2><p class="article-intro">{esc(opening)} {esc(profile['priority'])}</p></section>
      <section class="article-section article-local-feature-section"><h2>{esc(local)} {esc(section_title)}</h2><p>{esc(profile['student'])}이라면 {esc(profile['check'])}부터 확인합니다. {esc(profile['bridge'])}</p><div class="article-card-grid">{first_cards}</div></section>
      <section class="article-section article-local-feature-section"><h2>{esc(transition_title)}</h2><p>{esc(evidence_intro)} 권장 확인 흐름은 {esc(profile['rhythm'])}입니다.</p><div class="article-target-list">{plan_cards}</div></section>
      <section class="article-section article-local-feature-section"><h2>{esc(local)} 중3 수학의 다음 단계 판단 기준</h2><p>{esc(transition_basis)} {esc(profile['priority'])}</p><div class="article-card-grid">{second_cards}</div></section>
      <section class="article-section article-local-feature-section"><h2>{esc(weekly_title)}</h2><p>{esc(weekly_intro)} {esc(profile['check'])}</p><div class="article-target-list">{weekly_cards}</div></section>
      <section class="article-closing"><p>{esc(local)} 학부모는 {esc(closing)} {esc(profile['bridge'])}</p></section>
    </section>'''


def build_faqs(row: dict[str, str], profile: dict[str, str]) -> list[tuple[str, str]]:
    local = row_value(row, "근처 수업가능 동네")
    title = f"{local} 중3 수학학원"
    center = row_value(row, "센터명") or f"{local} 학습센터"
    schools = base.split_items(row_value(row, "타깃학교\n(중)"))
    grades = base.split_items(row_value(row, "가능학년\n(수학)"))
    school_answer = (
        f"공개 센터 자료에는 {', '.join(schools)} 등이 안내되어 있습니다. 실제 시험 범위와 일정은 학생이 가져온 학교 자료로 확인해야 합니다."
        if schools
        else "공개 센터 자료에는 중학교명이 별도로 기재되어 있지 않습니다. 학교를 임의로 단정하지 않고 상담에서 현재 학교 자료를 확인합니다."
    )
    availability_answer = (
        f"공개된 {center} 수학 가능 학년에 중3이 포함되어 있습니다. 실제 반 편성, 시간표와 시작 가능일은 상담에서 다시 확인해야 합니다."
        if "중3" in grades
        else f"공개된 {center} 자료에는 중3 수학 가능 여부가 기재되어 있지 않습니다. 개설 여부와 일정은 상담 확인이 필요합니다."
    )
    bank = [
        (f"{title} 상담 전에 무엇을 준비하면 좋나요?", f"최근 시험지, 현재 교재, 학교 진도와 누적 오답을 준비하세요. 특히 {profile['check']}를 메모하면 상담 기준이 구체적입니다."),
        (f"{local} 중3 수학학원은 어떤 학생에게 맞는지 어떻게 보나요?", f"{profile['student']}이라면 설명 뒤 독립 풀이와 간격 재풀이를 어떻게 확인하는지 살펴보세요. {profile['priority']}"),
        ("고교 선행과 중3 내신 중 무엇을 먼저 확인해야 하나요?", f"선행 범위를 일률적으로 앞당기기보다 현재 학교 진도와 누적 개념의 연결 상태를 먼저 봐야 합니다. {profile['bridge']}"),
        ("중3 수학의 서술형은 어떻게 점검하나요?", "정답뿐 아니라 주어진 조건, 사용한 개념, 식과 결론이 답안에 드러나는지 확인합니다. 이후 조건이 달라진 문제에서도 같은 근거를 사용할 수 있는지 봅니다."),
        ("학교 내신 자료는 어떤 기준으로 활용하나요?", school_answer),
        ("중3 수학 수강 가능 여부는 어디에서 확인하나요?", availability_answer),
        ("누적 공백이 있는 학생도 현재 진도를 이어갈 수 있나요?", "현재 단원을 막는 이전 개념을 찾아 짧게 복원하고 바로 현재 문제에 적용하는 방식으로 범위를 조정할 수 있습니다. 실제 순서는 진단 결과로 정합니다."),
        ("시험 범위가 정해지기 전에는 무엇을 복습하나요?", "학교 진도를 유지하면서 최근 오답과 이전 개념 공백을 정리합니다. 범위가 확정되면 교과서·학교 자료, 서술형과 시간 관리 중심으로 계획을 전환합니다."),
        ("오답은 얼마나 간격을 두고 다시 확인하나요?", f"수업 직후 정리에 그치지 않고 주중과 누적 확인 시점을 나누는 것이 좋습니다. {profile['rhythm']}이 실제 일정에 가능한지 확인하세요."),
        ("상담할 때 교습비와 운영 조건도 함께 확인해야 하나요?", "네. 교습비, 주당 횟수, 시작·종료 시각, 결석·보강 기준과 시험 기간 일정 변동을 같은 표에 적어 비교하는 것이 좋습니다."),
        ("시간 제한 문제는 언제 시작하는 것이 좋나요?", "풀이 순서와 정확도가 먼저 안정되어야 합니다. 이후 풀이 시간과 검산 시간을 따로 기록하고 학교 범위 안에서 완료 기준을 조정합니다."),
    ]
    required = [bank[0], bank[1], bank[4], bank[5]]
    ordered = sorted(bank, key=lambda item: base.seed_for(CATEGORY, local, "faq", item[0]))
    selected = required + [item for item in ordered if item not in required][:3]
    selected.sort(key=lambda item: base.seed_for(CATEGORY, local, "faq-order", item[0]))
    contexts = [
        f"{local}에서는 실제 풀이 기록과 주중 일정을 함께 놓고 이 기준을 확인하세요.",
        f"{center} 상담에서 이 항목이 다음 과제와 재풀이 날짜에 어떻게 반영되는지 질문하는 것이 좋습니다.",
        f"{profile['student']}이라면 문제 수보다 혼자 해결한 범위를 근거로 판단해야 합니다.",
        f"상담 전에 {profile['check']}를 적어 두면 관리 방식을 더 구체적으로 비교할 수 있습니다.",
        "학교 자료가 있다면 답만 보지 말고 풀이 흔적과 다시 푼 결과를 나란히 확인하세요.",
        "실제 적용 범위는 학교 일정, 학생 진단과 복습 가능 시간을 확인한 뒤 정해야 합니다.",
    ]
    return [(question, f"{answer} {choose(local, 'faq-context-' + question, contexts)}") for question, answer in selected]


def build_checklist(row: dict[str, str], profile: dict[str, str], page_index: int) -> list[tuple[str, str]]:
    local = row_value(row, "근처 수업가능 동네")
    center = row_value(row, "센터명") or f"{local} 학습센터"
    schools = base.split_items(row_value(row, "타깃학교\n(중)"))
    grades = base.split_items(row_value(row, "가능학년\n(수학)"))
    tuition = row_value(row, "센터 교습비")
    school = schools[base.seed_for(CATEGORY, local, "check-school") % len(schools)] if schools else "현재 학교 자료"
    checklist = [
        ("준비 자료", choose(local, "check-material", [f"{school}의 최근 시험지와 현재 교재의 풀이 흔적을 준비합니다.", f"{local} 상담 전에 교재 진도, 학교 범위와 누적 오답을 한곳에 모읍니다.", "최근 시험지, 오답노트와 질문 표시가 남은 현재 교재를 준비합니다."])),
        ("연결 공백", choose(local, "check-gap", [profile["check"], "현재 단원을 막는 중1·중2 개념을 한 가지로 좁힙니다.", "설명 뒤 혼자 다시 풀 수 없는 문제의 첫 단계를 표시합니다."])),
        ("주간 시간", choose(local, "check-time", [f"{profile['rhythm']} 흐름이 실제 일정에 가능한지 계산합니다.", "학교 진도와 누적 복습에 사용할 요일과 시간을 따로 적습니다.", f"{local}의 통학·귀가 시간을 제외한 실제 복습 시간을 확인합니다."])),
        ("운영 조건", choose(local, "check-operation", [f"{center}의 반 편성, 시작 가능일과 보강 기준을 상담에서 확인합니다.", f"{'공개 교습비 링크' if tuition else '상담 안내'}와 주당 횟수·시간표를 함께 비교합니다.", f"{'공개 자료의 중3 수학 가능 학년과' if '중3' in grades else '중3 수학 개설 여부와'} 실제 일정을 다시 확인합니다."])),
    ]
    unique_signals = [
        "독립 풀이", "오답 재현", "조건 해석", "검산 순서", "서술형 근거", "누적 개념",
        "주간 완료", "질문 기록", "시간 배분", "학교 자료", "복습 간격", "고교 연결",
        "풀이 설명", "교재 진도", "시험 계획", "운영 일정", "개념 적용", "재풀이 결과",
        "오류 분류", "완료 기준",
    ]
    unique_actions = [
        "최근 답안과 대조", "다음 확인 날짜 기록", "학교 진도와 분리", "주중 일정에 배치",
        "상담 질문으로 준비", "해설 없는 결과 확인", "우선순위 표에 반영", "시험 전후로 구분",
        "학생 설명과 비교", "완료 여부 재확인", "현재 단원에 적용", "누적 기록으로 보관",
        "보강 기준과 함께 확인", "풀이 흔적에 표시", "실행 가능한 분량 계산", "다음 과제와 연결",
        "검산 시간까지 기록", "서술형 답안에 적용", "교재 문제와 함께 점검", "상담 자료로 정리",
    ]
    signal = unique_signals[page_index % len(unique_signals)]
    action = unique_actions[(page_index // len(unique_signals)) % len(unique_actions)]
    label, body = checklist[-1]
    checklist[-1] = (label, f"{body} 이번 비교에서는 {signal} 항목을 {action}합니다.")
    return checklist


def build_parent_views(row: dict[str, str], profile: dict[str, str]) -> list[str]:
    local = row_value(row, "근처 수업가능 동네")
    center = row_value(row, "센터명") or f"{local} 학습센터"
    schools = base.split_items(row_value(row, "타깃학교\n(중)"))
    evidence = ", ".join(schools[:2]) if schools else "현재 학교 자료"
    options = [
        f"{local} 상담에서는 진도보다 학생이 틀린 이유를 설명하고 해설 없이 다시 풀 수 있는지 확인하는 편이 적절합니다.",
        f"{profile['student']}이라면 주간 복습 시점과 오답 완료 기록이 실제로 남는지 살펴볼 수 있습니다.",
        f"{center} 공개 정보는 수학 가능 학년, 교습비 확인 경로와 실제 운영 조건을 구분해 보는 것이 좋습니다.",
        f"{evidence}를 준비하면 개념 부족, 조건 해석과 계산 실수를 실제 답안으로 구분할 수 있습니다.",
        "고교 선행 범위보다 중학교 개념을 새 문제에 독립적으로 적용하는지 먼저 확인하는 것이 안전합니다.",
        "시험 범위 확정 전에는 누적 공백을, 확정 후에는 학교 자료와 오답의 완료 순서를 따로 확인해야 합니다.",
        f"{profile['check']}를 상담 질문으로 적어 두면 문제량보다 관리 방식의 차이를 구체적으로 비교할 수 있습니다.",
        f"{local}에서는 종료 시각, 결석·보강 기준과 시험 기간 시간표가 실제 주간 일정에 맞는지도 살펴봐야 합니다.",
        "서술형은 식과 답뿐 아니라 사용한 개념과 조건, 결론 사이의 근거가 남는지 확인할 필요가 있습니다.",
        "시간 제한은 정확한 풀이 순서가 자리 잡은 뒤 적용하는지, 검산 시간이 별도로 확보되는지 확인해야 합니다.",
    ]
    options.sort(key=lambda value: base.seed_for(CATEGORY, local, "parent-view", value))
    return options[:3]


def page_json_ld(
    row: dict[str, str],
    description: str,
    page_url: str,
    rep_image: str,
    body_image: str,
    map_image: str,
    faqs: list[tuple[str, str]],
    related: list[tuple[str, str]],
) -> dict:
    local = row_value(row, "근처 수업가능 동네")
    region = row_value(row, "지역")
    district = row_value(row, "시or구")
    title = f"{local} 중3 수학학원"
    center = row_value(row, "센터명") or f"{local} 학습센터"
    address = row_value(row, "센터 주소")
    registration = row_value(row, "교육지원청 등록번호")
    schools = base.split_items(row_value(row, "타깃학교\n(중)"))
    grades = base.split_items(row_value(row, "가능학년\n(수학)"))
    tuition = row_value(row, "센터 교습비")
    organization = {
        "@type": ["EducationalOrganization", "LocalBusiness"],
        "@id": f"{page_url}#organization",
        "name": center,
        "url": page_url,
        "telephone": base.PHONE_DISPLAY,
        "areaServed": {"@type": "Place", "name": f"{region} {district} {local}"},
        "address": {"@type": "PostalAddress", "streetAddress": address, "addressCountry": "KR"},
        "educationalLevel": grades,
        "knowsAbout": ["중3 수학", "중학교 내신", "서술형", "누적 개념", "고교 진입 전 학습", "시간 관리"],
        "contactPoint": {"@type": "ContactPoint", "telephone": f"+82-{base.PHONE_DISPLAY[1:]}", "contactType": "교육 상담", "availableLanguage": "Korean", "url": base.FORM_URL},
        "makesOffer": [{"@type": "Offer", "name": f"{title} 학습 상담", "category": "중3 수학 학습 진단", "url": tuition or page_url, "itemOffered": {"@id": f"{page_url}#service"}}],
        "mentions": [{"@type": "School", "name": school} for school in schools],
    }
    if registration:
        organization["identifier"] = registration
    return {
        "@context": "https://schema.org",
        "@graph": [
            {"@type": "WebSite", "@id": f"{DOMAIN}/#website", "url": f"{DOMAIN}/", "name": SITE_NAME, "inLanguage": "ko-KR"},
            organization,
            {"@type": "WebPage", "@id": f"{page_url}#webpage", "url": page_url, "name": title, "description": description, "inLanguage": "ko-KR", "isPartOf": {"@id": f"{DOMAIN}/#website"}, "breadcrumb": {"@id": f"{page_url}#breadcrumb"}, "mainEntity": {"@id": f"{page_url}#service"}, "primaryImageOfPage": {"@id": f"{page_url}#primaryimage"}, "about": [{"@type": "Place", "name": f"{region} {district} {local}"}, {"@type": "Thing", "name": "중3 수학학원"}, {"@type": "Thing", "name": "고교 진입 전 수학 학습"}], "mentions": [{"@type": "EducationalOrganization", "name": center}, *[{"@type": "School", "name": school} for school in schools]], "hasPart": [{"@type": "WebPageElement", "name": name} for name in ["핵심 답변", "중3 수학 학습 설계", "센터 정보", "상담 전 체크리스트", "FAQ", "학부모 상담 관점", "관련 페이지"]]},
            {"@type": "ImageObject", "@id": f"{page_url}#primaryimage", "url": rep_image, "caption": f"{title} {SITE_NAME} 대표"},
            {"@type": "BreadcrumbList", "@id": f"{page_url}#breadcrumb", "itemListElement": [{"@type": "ListItem", "position": 1, "name": "홈", "item": f"{DOMAIN}/"}, {"@type": "ListItem", "position": 2, "name": PARENT, "item": canonical(PARENT)}, {"@type": "ListItem", "position": 3, "name": CATEGORY_LABEL, "item": canonical(PARENT, CATEGORY)}, {"@type": "ListItem", "position": 4, "name": title, "item": page_url}]},
            {"@type": "Service", "@id": f"{page_url}#service", "name": f"{title} 학습 상담 및 안내", "serviceType": "중학교 3학년 수학 학습관리", "provider": {"@id": f"{page_url}#organization"}, "areaServed": {"@type": "Place", "name": f"{region} {district} {local}"}, "audience": {"@type": "EducationalAudience", "educationalRole": "중학교 3학년 학생 및 학부모"}, "about": ["중3 수학", "내신", "서술형", "누적 공백", "고교 준비", "시간 관리"]},
            {"@type": "Article", "@id": f"{page_url}#article", "url": page_url, "headline": title, "description": description, "datePublished": PUBLISH_DATE, "dateModified": PUBLISH_DATE, "inLanguage": "ko-KR", "mainEntityOfPage": {"@id": f"{page_url}#webpage"}, "author": {"@id": f"{page_url}#organization"}, "publisher": {"@id": f"{page_url}#organization"}, "image": [rep_image, body_image, map_image], "articleSection": [region, district, local, "중3 수학", "내신", "서술형", "고교 진입 전"], "about": [{"@type": "Thing", "name": title}], "mentions": [{"@type": "EducationalOrganization", "name": center}]},
            {"@type": "FAQPage", "@id": f"{page_url}#faq", "mainEntity": [{"@type": "Question", "name": question, "acceptedAnswer": {"@type": "Answer", "text": answer}} for question, answer in faqs]},
            {"@type": "ItemList", "@id": f"{page_url}#related-pages", "name": f"{local} 중3 수학 관련 페이지", "itemListElement": [{"@type": "ListItem", "position": index, "name": name, "url": url} for index, (name, url) in enumerate(related, 1)]},
        ],
    }


def detail_html(
    row: dict[str, str],
    manuscript: str,
    repeated_signatures: set[str],
    image_row: dict[str, str],
    rep_image: str,
    map_name: str,
    peer_locals: list[str],
    page_index: int,
) -> str:
    local = row_value(row, "근처 수업가능 동네")
    region = row_value(row, "지역")
    district = row_value(row, "시or구")
    center = row_value(row, "센터명") or f"{local} 학습센터"
    address = row_value(row, "센터 주소")
    location = base.inline_text(row_value(row, "위치안내"))
    registration_name = row_value(row, "교육지원청명칭")
    registration_number = row_value(row, "교육지원청 등록번호")
    tuition = row_value(row, "센터 교습비")
    grades = base.split_items(row_value(row, "가능학년\n(수학)"))
    schools = base.split_items(row_value(row, "타깃학교\n(중)"))
    available = "중3" in grades
    title = f"{local} 중3 수학학원"
    profile = page_profile(row)
    description = build_description(row, profile, page_index)
    page_url = canonical(PARENT, CATEGORY, local)
    body_name = "seoul.jpg" if region == "서울" else "local.jpg"
    body_url = f"{DOMAIN}/assets/centers/common/{body_name}"
    map_url = f"{DOMAIN}/assets/maps/{quote(map_name)}"
    faqs = build_faqs(row, profile)
    checklist = build_checklist(row, profile, page_index)
    parent_views = build_parent_views(row, profile)
    related = [(CATEGORY_LABEL, canonical(PARENT, CATEGORY)), (f"{local} 중2 수학학원", canonical(PARENT, "중2수학학원", local))]
    related.extend((f"{peer} 중3 수학학원", canonical(PARENT, CATEGORY, peer)) for peer in peer_locals)
    schema = page_json_ld(row, description, page_url, rep_image, body_url, map_url, faqs, related)
    manuscript_html = contextualize_manuscript(manuscript, row, repeated_signatures)
    article = build_learning_article(row, profile, page_index)
    school_html = "".join(f"<span>{esc(school)}</span>" for school in schools) if schools else '<p class="subject-empty-note">공개 자료에 중학교명이 별도로 기재되어 있지 않아 상담에서 현재 학교 자료를 확인합니다.</p>'
    grade_html = "".join(f"<span>{esc(grade)}</span>" for grade in grades) if grades else "<span>상담 확인</span>"
    location_html = f'<article data-role="verified-location"><span>확인된 위치 안내</span><strong>{esc(center)}</strong><p>{esc(location)}</p></article>' if location else ""
    tuition_html = f'<a class="btn ghost" href="{esc(tuition)}" target="_blank" rel="noopener noreferrer">센터 교습비 자료 확인</a>' if tuition else '<span class="subject-empty-note">교습비 자료는 상담에서 확인해 주세요.</span>'
    checklist_html = "".join(f'<article class="geo-check-card"><b>{index:02d}</b><strong>{esc(label)}</strong><p>{esc(body)}</p></article>' for index, (label, body) in enumerate(checklist, 1))
    faq_html = "".join(f"<details><summary>{esc(question)}</summary><p>{esc(answer)}</p></details>" for question, answer in faqs)
    parent_html = "".join(f'<article class="subject-parent-note"><p>{esc(note)}</p></article>' for note in parent_views)
    peers_html = "".join(f'<a class="child-page-button" href="../{quote(peer, safe="")}/index.html">{esc(peer)} 중3 수학학원</a>' for peer in peer_locals)
    return f'''{base.head_html(title, description, page_url, rep_image, schema, "../../../")}
<body>
{base.nav_html("../../../", "subject")}
  <main id="main">
    <nav class="breadcrumb-box" aria-label="현재 위치"><a href="../../../index.html">홈</a><span>›</span><a href="../../index.html">과목별학원</a><span>›</span><a href="../index.html">중3 수학학원</a><span>›</span>{esc(title)}</nav>
    <section class="sub-hero shell directory-hero subject-detail-hero"><div class="reveal"><p class="eyebrow">MIDDLE 3 MATH GUIDE</p><h1>{esc(title)}</h1><p>{esc(description)}</p><div class="hero-actions"><a class="btn primary" href="{base.FORM_URL}" target="_blank" rel="noopener noreferrer">상담 준비하기</a><a class="btn ghost" href="../index.html">다른 지역 찾기</a></div></div><div class="stat-console reveal"><div class="stat-pill"><strong>{esc(local)}</strong><span>{esc(region)} {esc(district)}</span></div><div class="stat-pill"><strong>{'중3' if available else '확인'}</strong><span>{'수학 가능 학년 기재' if available else '개설 여부 상담 필요'}</span></div></div></section>
    <section class="shell csv-body-stack csv-top-media local-media-section subject-media" aria-label="{esc(title)} 이미지 안내"><img data-role="representative-image" style="display:none;" src="{esc(rep_image)}" alt="{esc(title)} {SITE_NAME} 대표"><figure class="csv-media-card"><img src="../../../assets/centers/common/{body_name}" alt="{esc(title)} 본문 {SITE_NAME}" loading="eager" decoding="async"></figure><figure class="csv-media-card"><img src="../../../assets/maps/{quote(map_name)}" alt="{esc(title)} 지도 {SITE_NAME}" loading="lazy" decoding="async"></figure></section>
    <section class="shell geo-summary-panel subject-answer-panel reveal" aria-labelledby="answer-title"><p class="eyebrow">핵심 답변</p><h2 id="answer-title">{esc(title)}, 무엇부터 확인해야 할까요?</h2><p>{esc(local)}에서 중3 수학학원을 비교할 때는 선행 범위만 보지 말고 현재 학교 진도, 중1·중2 누적 공백, 서술형 근거, 오답 재확인과 시간 관리 기록을 함께 살펴야 합니다.</p><div class="geo-fact-grid"><article class="geo-fact-card"><span>현재 학생 상황</span><strong>{esc(profile['student'])}</strong></article><article class="geo-fact-card"><span>우선 확인</span><strong>{esc(profile['priority'])}</strong></article><article class="geo-fact-card"><span>고교 연결</span><strong>{esc(profile['bridge'])}</strong></article></div></section>
    {manuscript_html}
    {article}
    <section class="shell academy-section subject-local-facts reveal" aria-labelledby="local-facts-title"><div class="section-heading"><p class="eyebrow">VERIFIED LOCAL FACTS</p><h2 id="local-facts-title">{esc(local)} 센터 정보와 확인 항목</h2><p>센터정보 자료에 기재된 내용만 사용했으며 확인되지 않은 학교와 운영 조건은 임의로 만들지 않았습니다.</p></div><div class="subject-fact-grid"><article><span>센터</span><strong>{esc(center)}</strong><p>{esc(address) if address else '주소는 상담에서 확인해 주세요.'}</p></article><article><span>중3 수학 가능 학년</span><strong>{'자료에 기재됨' if available else '상담 확인 필요'}</strong><p>{'공개 자료에 중3 수학이 포함되어 있으며 실제 반 편성과 일정은 상담에서 확인합니다.' if available else '공개 자료에 중3 수학 가능 여부가 기재되어 있지 않아 상담 확인이 필요합니다.'}</p></article><article><span>교육지원청 등록 정보</span><strong>{esc(registration_name) if registration_name else '상담 확인'}</strong><p>{esc(registration_number) if registration_number else '공개 자료에 등록번호가 별도로 기재되어 있지 않습니다.'}</p></article><article><span>교습비 확인</span><strong>{'공개 링크 확인 가능' if tuition else '상담 확인 필요'}</strong><p>주당 횟수·시간표·보강 기준과 함께 비교합니다.</p></article>{location_html}</div><div class="subject-school-panel"><h3>공개 자료의 중학교 안내 · {len(schools)}개</h3><div class="subject-school-tags">{school_html}</div></div><div class="subject-grade-panel"><h3>공개 자료의 수학 가능 학년</h3><div class="subject-school-tags">{grade_html}</div>{tuition_html}</div></section>
    <section class="shell geo-checklist-panel reveal" aria-labelledby="checklist-title"><p class="eyebrow">상담 전 체크리스트</p><h2 id="checklist-title">{esc(title)} 비교 전에 적어둘 내용</h2><div class="geo-checklist-grid">{checklist_html}</div></section>
    <section class="shell academy-section local-proof-section" aria-labelledby="faq-title"><div class="section-heading"><p class="eyebrow">FAQ & PARENT VIEW</p><h2 id="faq-title">{esc(title)} 자주 묻는 질문과 학부모 상담 관점</h2><p>화면 질문과 답변은 JSON-LD에도 동일하게 반영했습니다. 상담 관점은 실제 후기를 가장하지 않고 학부모가 확인할 비교 기준으로 정리했습니다.</p></div><div class="local-proof-layout"><section class="local-faq-card" aria-label="{esc(title)} 자주 묻는 질문"><div class="faq-list">{faq_html}</div></section><aside class="local-review-card" aria-label="{esc(title)} 학부모 상담 관점"><div class="review-list">{parent_html}</div></aside></div></section>
    <section class="shell local-page-nav reveal" aria-labelledby="related-title"><div class="section-heading"><p class="eyebrow">RELATED GUIDES</p><h2 id="related-title">{esc(local)} 중2·중3 수학과 인접 지역 비교</h2><p>같은 동네 중2 수학 안내와 시군구·광역권을 우선한 중3 수학 페이지를 연결했습니다.</p></div><div class="child-button-grid"><a class="child-page-button" href="../index.html">중3 수학학원 지역 목록</a><a class="child-page-button" href="../../중2수학학원/{quote(local, safe='')}/index.html">{esc(local)} 중2 수학학원</a>{peers_html}</div></section>
  </main>
{base.footer_html("../../../")}
</body>
</html>
'''


def hub_page(rows: list[dict[str, str]]) -> str:
    title = "중3 수학학원 지역별 안내"
    description = "371개 지역의 중3 수학학원 선택 기준과 내신·서술형·누적 공백·고교 진입 전 학습·시간 관리 확인 항목을 정리했습니다."
    url = canonical(PARENT, CATEGORY)
    items = [(f"{row_value(row, '근처 수업가능 동네')} 중3 수학학원", canonical(PARENT, CATEGORY, row_value(row, "근처 수업가능 동네"))) for row in rows]
    faqs = [
        ("지역별 중3 수학학원 페이지는 어떤 자료로 구성했나요?", "센터정보 정리 자료의 지역, 센터명, 주소, 실제 기재 학교, 수학 가능 학년과 교습비 링크를 기준으로 구성했습니다."),
        ("중3 수학 가능 학년이 확인되지 않는 지역도 있나요?", "네. 공개 자료에 중3 수학이 기재되지 않은 지역은 개설을 임의로 단정하지 않고 상담 확인 필요로 표시합니다."),
        ("고교 준비는 어떤 기준으로 안내하나요?", "선행 범위를 약속하지 않고 중학교 개념 연결, 독립 풀이, 서술형 근거, 누적 공백과 시간 관리 상태를 확인하는 기준을 안내합니다."),
    ]
    schema = base.collection_schema(title, description, url, [("홈", f"{DOMAIN}/"), (PARENT, canonical(PARENT)), (CATEGORY_LABEL, url)], items, faqs)
    grouped: dict[str, dict[str, list[dict[str, str]]]] = defaultdict(lambda: defaultdict(list))
    for row in rows:
        grouped[row_value(row, "지역")][row_value(row, "시or구")].append(row)
    region_sections = []
    for region_index, (region, districts) in enumerate(grouped.items()):
        district_sections = []
        for district, district_rows in districts.items():
            cards = []
            for row in district_rows:
                local = row_value(row, "근처 수업가능 동네")
                search = " ".join([region, district, local, row_value(row, "센터명"), row_value(row, "타깃학교\n(중)")])
                cards.append(f'<a class="subject-town-card" data-subject-town data-search="{esc(search)}" href="{quote(local, safe="")}/index.html"><strong>{esc(local)}</strong><span>{esc(district)} · 중3 수학</span></a>')
            district_sections.append(f'<section class="subject-district-group" data-subject-district><h3>{esc(district)} <small>{len(district_rows)}개 지역</small></h3><div class="subject-town-grid">{"".join(cards)}</div></section>')
        count = sum(len(values) for values in districts.values())
        region_sections.append(f'<details class="subject-region-group" data-subject-region{" open" if region_index == 0 else ""}><summary><span>{esc(region)}</span><b>{count}개 지역</b></summary><div class="subject-region-body">{"".join(district_sections)}</div></details>')
    faq_html = "".join(f"<details><summary>{esc(question)}</summary><p>{esc(answer)}</p></details>" for question, answer in faqs)
    return f'''{base.head_html(title, description, url, f"{DOMAIN}/assets/generated/academy-hero-v2.webp", schema, "../../", "website")}
<body>
{base.nav_html("../../", "subject")}
  <main id="main"><nav class="breadcrumb-box" aria-label="현재 위치"><a href="../../index.html">홈</a><span>›</span><a href="../index.html">과목별학원</a><span>›</span>중3 수학학원</nav><section class="sub-hero shell directory-hero"><div class="reveal"><p class="eyebrow">MIDDLE 3 MATH DIRECTORY</p><h1>중3 수학학원</h1><p>{esc(description)}</p></div><div class="stat-console reveal"><div class="stat-pill"><strong>371</strong><span>동네별 학습 안내</span></div><div class="stat-pill"><strong>13</strong><span>광역 지역 구분</span></div></div></section><section class="shell academy-section subject-directory" data-subject-directory><div class="section-heading"><p class="eyebrow">LOCAL SEARCH</p><h2>동네 이름으로 중3 수학학원 찾기</h2><p>광역·시군구별로 접어 정리했습니다. 동네명이나 센터명을 검색하면 해당 결과만 남습니다.</p></div><div class="subject-search-box"><label for="subject-town-search">동네·시군구·센터 검색</label><input id="subject-town-search" type="search" placeholder="예: 명일동, 강동구, 명일점" autocomplete="off" data-subject-search><p aria-live="polite" data-subject-search-status>전체 371개 지역</p></div><div class="subject-region-list">{"".join(region_sections)}</div></section><section class="shell academy-section subject-hub-faq"><div class="section-heading"><p class="eyebrow">DIRECTORY FAQ</p><h2>지역 페이지 이용 전에 확인하세요</h2></div><div class="faq-list">{faq_html}</div></section></main>
{base.footer_html("../../")}
</body>
</html>
'''


def main() -> None:
    rows = base.read_csv(base.CENTER_CSV)
    manuscripts = load_manuscripts()
    if len(rows) != 371:
        raise ValueError(f"센터 자료가 371개가 아닙니다: {len(rows)}")
    if len(manuscripts) != 371:
        raise ValueError(f"중3 수학 원고가 371개가 아닙니다: {len(manuscripts)}")
    sanitized_manuscripts = [
        sanitize_middle3_manuscript(manuscript, row)
        for manuscript, row in zip(manuscripts, rows)
    ]
    repeated_signatures = base.repeated_paragraphs(sanitized_manuscripts, rows)
    locals_ = [row_value(row, "근처 수업가능 동네") for row in rows]
    if len(set(locals_)) != 371:
        raise ValueError("동네명이 중복되었습니다.")
    image_rows = base.load_image_rows()
    representatives = base.representative_urls()
    if not representatives:
        raise ValueError("대표 이미지 URL을 찾지 못했습니다.")
    peer_network = base.build_subject_peer_network(rows)
    if any(len(peers) != 6 for peers in peer_network.values()):
        raise ValueError("중3 수학 상호 형제 링크가 6개가 아닙니다.")
    target = SITE / PARENT / CATEGORY
    target.mkdir(parents=True, exist_ok=True)
    (SITE / PARENT / "index.html").write_text(
        base.clean_markup(base.category_page()), encoding="utf-8", newline="\n"
    )
    (target / "index.html").write_text(
        base.clean_markup(hub_page(rows)), encoding="utf-8", newline="\n"
    )
    sitemap_urls = [
        (canonical(PARENT), "weekly", "0.9"),
        (canonical(PARENT, CATEGORY), "weekly", "0.9"),
    ]
    for page_index, (row, manuscript) in enumerate(zip(rows, manuscripts)):
        local = row_value(row, "근처 수업가능 동네")
        image_row = image_rows.get(local, {})
        folder = target / local
        folder.mkdir(parents=True, exist_ok=True)
        page = detail_html(
            row,
            manuscript,
            repeated_signatures,
            image_row,
            base.representative_for(row, representatives),
            base.map_filename(row, image_row),
            peer_network[local],
            page_index,
        )
        (folder / "index.html").write_text(
            base.clean_markup(page), encoding="utf-8", newline="\n"
        )
        sitemap_urls.append((canonical(PARENT, CATEGORY, local), "monthly", "0.8"))
    base.update_sitemap(sitemap_urls)
    print(f"generated_hub=1")
    print(f"generated_details={len(rows)}")
    print(f"generated_total={len(rows) + 1}")
    print(f"target={target}")
    print("category_index_updated=true")
    print(f"sitemap_urls_checked={len(sitemap_urls)}")


if __name__ == "__main__":
    main()
