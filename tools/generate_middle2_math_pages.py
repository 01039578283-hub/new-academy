from __future__ import annotations

import csv
import hashlib
import html
import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from urllib.parse import quote

from openpyxl import load_workbook


SITE = Path(__file__).resolve().parents[1]
REFERENCE = SITE.parent / "참고자료"
COMMON = REFERENCE / "공통자료"
MANUSCRIPT = REFERENCE / "원고모음(엑셀)" / "중2 수학학원 원고.xlsx"
CENTER_CSV = COMMON / "센터정보 정리.csv"
IMAGE_CSV = COMMON / "이미지링크.csv"
REPRESENTATIVE_CSV = COMMON / "대표 이미지 url.csv"

DOMAIN = "https://xn--z92bu9jx8cwzc.com"
SITE_NAME = "와와센터"
PHONE_DISPLAY = "010-3957-8283"
PHONE_LINK = "01039578283"
FORM_URL = "https://docs.google.com/forms/d/e/1FAIpQLSdb2oE5Qk5YS0TfYDxyV1w-IOTkhkjOCmmpAKTI9FmqpVj6Yg/viewform"
SMS_URL = "https://blogsms.net/01039578283"
PUBLISH_DATE = "2026-08-06"

PARENT = "과목별학원"
CATEGORY = "중2수학학원"
CATEGORY_LABEL = "중2 수학학원"


def esc(value: object) -> str:
    return html.escape(str(value or ""), quote=True)


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def row_value(row: dict[str, str], needle: str) -> str:
    for key, value in row.items():
        if needle in key:
            return (value or "").strip()
    return ""


def split_items(value: str) -> list[str]:
    return [item.strip() for item in re.split(r"[,/|]", value or "") if item.strip()]


def inline_text(value: str) -> str:
    return re.sub(r"\s+", " ", value or "").strip()


def object_form(value: str) -> str:
    value = (value or "").strip()
    if not value:
        return value
    code = ord(value[-1]) - 0xAC00
    has_final = 0 <= code <= 11171 and code % 28 != 0
    return f"{value}{'을' if has_final else '를'}"


def clean_markup(value: str) -> str:
    return "\n".join(line.rstrip() for line in value.splitlines()).rstrip() + "\n"


def seed_for(*parts: str) -> int:
    digest = hashlib.sha256("|".join(parts).encode("utf-8")).digest()
    return int.from_bytes(digest[:8], "big")


def choose(local: str, label: str, values: list[str]) -> str:
    return values[seed_for(CATEGORY, local, label) % len(values)]


def canonical(*parts: str) -> str:
    path = "/" + "/".join(part.strip("/") for part in parts if part) + "/"
    return DOMAIN + quote(path, safe="/")


def load_manuscripts() -> list[str]:
    if not MANUSCRIPT.exists():
        raise FileNotFoundError(MANUSCRIPT)
    workbook = load_workbook(MANUSCRIPT, read_only=True, data_only=True)
    sheet = workbook.active
    values = [str(row[0] or "").strip() for row in sheet.iter_rows(values_only=True)]
    workbook.close()
    return values


def strip_tags(value: str) -> str:
    return re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", value)).strip()


def paragraph_signature(value: str, local: str) -> str:
    text = strip_tags(value).replace(local, "<LOCAL>")
    text = re.sub(r"\d+(?:[-~–]\d+)*", "<NUM>", text)
    return text


def repeated_paragraphs(manuscripts: list[str], rows: list[dict[str, str]]) -> set[str]:
    counts: Counter[str] = Counter()
    for raw, row in zip(manuscripts, rows):
        local = row_value(row, "근처 수업가능 동네")
        for body in re.findall(r"<p(?:\s[^>]*)?>(.*?)</p>", raw, re.DOTALL | re.I):
            signature = paragraph_signature(body, local)
            if len(signature) >= 35:
                counts[signature] += 1
    return {signature for signature, count in counts.items() if count > 1}


def sanitize_manuscript(value: str, row: dict[str, str]) -> str:
    """수학 검색 의도와 확인 가능한 표현만 남기도록 외부 원고를 정제합니다."""
    local = row_value(row, "근처 수업가능 동네")
    replacements = {
        "바로 성적이 오르는 방식으로": "학습 과정을 점검하는 방식으로",
        "시험 점수로 연결되도록": "시험 대비 흐름에 반영되도록",
        "점수로 연결되도록": "시험 준비에 반영되도록",
        "다시는 반복되지 않게": "같은 실수가 줄어들도록",
        "실력을 완성합니다": "실력을 단계적으로 정리합니다",
        "성적 향상을": "학습 개선을",
        "영어·수학·국어·영수까지 함께 고민하는 학생": "중2 수학의 개념·유형·서술형·오답을 함께 점검하려는 학생",
        "영어·국어·영수 함께 잡는 코칭": "수학 개념과 풀이 과정을 함께 확인하는 코칭",
        "영어·국어·영수까지 학습 로드맵": "수학 개념·풀이·오답까지의 학습 로드맵",
        "수학만 올리는 방식이 아니라": "문제량만 늘리는 방식이 아니라",
        "수학+기초 과목 연계 관리": "수학 개념과 조건 해석 연결 관리",
        "영수(통합 관리)": "수학 학습관리",
        "성적과 문제 해결력을 함께 올릴 수 있도록": "개념과 문제 해결 과정을 함께 점검하도록",
        "시험 직전까지 안정적인 점수를 만듭니다": "시험 전까지 확인할 오답과 완료 범위를 구분합니다",
        "다음 점수로 바꿉니다": "다음 풀이의 기준으로 바꿉니다",
        "시험 당일 점수로 연결되게": "시험 준비 과정에 반영되게",
        "내신에서 점수로 이어지도록": "내신 준비 과정에 반영되도록",
        "성적 상승을 돕습니다": "학습 개선을 위한 실행 기준을 정리합니다",
        "성적 상승을 노립니다": "학습 개선을 위한 실행 기준을 정리합니다",
        "성적 상승의 기반을 만듭니다": "학습 개선을 위한 실행 기준을 정리합니다",
        "점수를 끌어올립니다": "재풀이 순서를 구체적으로 정합니다",
        "속도와 정확도를 동시에 끌어올립니다": "풀이 시간과 검산 정확도를 함께 점검합니다",
        "빈출 유형부터": "확인된 학교 자료의 유형부터",
        "시험에 나오는 출제 흐름에 맞춰": "현재 시험 범위의 유형을 기본·적용·재확인 순서로 나누어",
        "지역 학생들의 내신 출제 흐름에 맞춰": "현재 학교 자료에서 확인되는 범위와 학생의 오답을 기준으로",
    }
    for before, after in replacements.items():
        value = value.replace(before, after)

    math_cards = [
        ("조건을 식으로 옮기는 과정", "문제의 조건을 표시하고 사용할 개념을 고른 뒤 식을 세우는 순서를 확인합니다."),
        ("풀이 근거를 남기는 연습", "정답만 확인하지 않고 사용한 개념과 계산 순서를 학생이 말하고 쓰게 합니다."),
        ("계산과 검산 기준", "부호·괄호·계산 순서와 답의 범위를 나누어 같은 기준으로 검산합니다."),
        ("오답 원인별 재풀이", "개념 부족·조건 해석·계산·풀이 순서 중 원인을 구분하고 간격을 두어 다시 풉니다."),
        ("학교 진도와 누적 복습", "현재 학교 자료와 교재 진도를 확인한 뒤 지금 단원을 막는 이전 개념부터 복원합니다."),
        ("서술형 답안 점검", "식과 답뿐 아니라 조건, 사용한 성질과 결론이 답안에 드러나는지 살펴봅니다."),
        ("풀이 시간 기록", "정확한 풀이 순서가 자리 잡은 뒤 문제별 시간과 검산 시간을 나누어 기록합니다."),
        ("독립 풀이 확인", "설명을 들은 직후와 며칠 뒤의 재풀이를 비교해 혼자 해결할 수 있는 범위를 확인합니다."),
    ]
    math_paragraphs = [
        "중2 수학은 현재 진도만 앞서가기보다 개념 이해, 조건 해석, 계산 과정과 재풀이 결과를 함께 확인해야 합니다.",
        "문제량보다 학생이 어느 단계에서 멈췄는지를 기록하고, 필요한 개념과 다음 확인 날짜를 구분하는 과정이 중요합니다.",
        "학교 진도와 최근 오답을 한 표에 놓고 기본 유형, 적용 문제와 서술형의 완료 기준을 따로 정합니다.",
        "수업에서 이해한 풀이를 며칠 뒤 해설 없이 다시 구성할 수 있는지 확인해 다음 학습 범위를 조정합니다.",
        "계산 실수와 개념 부족을 같은 오답으로 묶지 않고 조건 해석, 식 세우기, 계산과 검산 단계로 나누어 살펴봅니다.",
        "현재 단원을 막는 이전 개념만 골라 짧게 복원한 뒤 학교 자료의 새 문제에 같은 기준을 적용합니다.",
        "시험 범위가 정해지기 전에는 누적 공백과 오답을 정리하고, 범위가 확정되면 교과서와 학교 자료 중심으로 계획을 조정합니다.",
        "정답을 맞힌 문제도 풀이 근거를 설명할 수 있는지 확인해 외운 풀이와 스스로 이해한 풀이를 구분합니다.",
        "주간 계획에는 문제 수뿐 아니라 막힌 문제, 질문할 내용, 재풀이 날짜와 완료 여부가 함께 기록되어야 합니다.",
        "서술형에서는 사용한 개념과 주어진 조건, 계산 결과가 자연스럽게 이어지는지 단계별로 점검합니다.",
        "시간 제한은 정확한 풀이 순서가 안정된 뒤 적용하고, 문제 풀이 시간과 검산 시간을 따로 확인합니다.",
        "상담에서는 최근 시험지와 교재의 풀이 흔적을 근거로 우선 보완할 내용과 유지할 내용을 나누는 것이 좋습니다.",
    ]

    cross_subject = re.compile(r"(?:영어|국어|영수)", re.I)

    def replace_cross_article(match: re.Match[str]) -> str:
        block = match.group(0)
        if not cross_subject.search(strip_tags(block)):
            return block
        opening_match = re.match(r"^(<article\b[^>]*>)", block, re.I | re.S)
        opening = opening_match.group(1) if opening_match else '<article class="article-card">'
        heading_tag = "h3" if re.search(r"<h3\b", block, re.I) else "strong"
        heading, body = math_cards[seed_for(CATEGORY, local, "sanitize-card", strip_tags(block)) % len(math_cards)]
        return f"{opening}<{heading_tag}>{esc(heading)}</{heading_tag}><p>{esc(body)}</p></article>"

    value = re.sub(
        r"<article\b[^>]*>.*?</article>",
        replace_cross_article,
        value,
        flags=re.I | re.S,
    )

    def replace_cross_text(match: re.Match[str]) -> str:
        tag, attrs_, body = match.group(1), match.group(2) or "", match.group(3)
        if not cross_subject.search(strip_tags(body)):
            return match.group(0)
        replacement = math_paragraphs[
            seed_for(CATEGORY, local, "sanitize-text", tag, strip_tags(body)) % len(math_paragraphs)
        ]
        return f"<{tag}{attrs_}>{esc(replacement)}</{tag}>"

    value = re.sub(
        r"<(p|li)(\s[^>]*)?>(.*?)</\1>",
        replace_cross_text,
        value,
        flags=re.I | re.S,
    )

    def replace_cross_heading(match: re.Match[str]) -> str:
        tag, attrs_, body = match.group(1), match.group(2) or "", match.group(3)
        if not cross_subject.search(strip_tags(body)):
            return match.group(0)
        heading, _body = math_cards[
            seed_for(CATEGORY, local, "sanitize-heading", strip_tags(body)) % len(math_cards)
        ]
        return f"<{tag}{attrs_}>{esc(heading)}</{tag}>"

    value = re.sub(
        r"<(h2|h3|strong)(\s[^>]*)?>(.*?)</\1>",
        replace_cross_heading,
        value,
        flags=re.I | re.S,
    )

    def replace_result_claim(match: re.Match[str]) -> str:
        options = [
            "학습 과정을 기록하고 다음 재풀이 기준을 정합니다.",
            "풀이 정확도와 검산 과정을 나누어 확인합니다.",
            "해결한 오답과 다시 확인할 오답을 구분해 계획에 반영합니다.",
            "특정 결과를 단정하지 않고 현재 풀이 기록을 기준으로 학습 범위를 조정합니다.",
            "학생이 혼자 해결할 수 있는 문제 범위와 다음 확인 날짜를 정리합니다.",
            "시험 전까지 확인할 단원과 오답의 완료 기준을 구체적으로 세웁니다.",
        ]
        original = match.group(0)
        leading = " " if original[:1].isspace() else ""
        return leading + options[seed_for(CATEGORY, local, "sanitize-result", original) % len(options)]

    value = re.sub(
        r"[^<>.!?]*(?:(?:성적|점수)[^<>.!?]{0,35}(?:상승|향상|올리|끌어올리|연결|만들|보장)|(?:상승|향상|올리|끌어올리)[^<>.!?]{0,20}(?:성적|점수)|보장)[^<>.!?]*(?:[.!?]|$)",
        replace_result_claim,
        value,
        flags=re.I,
    )

    def replace_school_claim(match: re.Match[str]) -> str:
        options = [
            "학생이 가져온 시험 범위·교과서·학교 프린트를 확인한 뒤 복습 순서를 정합니다.",
            "현재 학교 자료에서 확인되는 범위와 학생의 오답을 기준으로 학습 순서를 조정합니다.",
            "학교별 일정을 임의로 단정하지 않고 실제 시험 범위와 교재 진도를 상담에서 확인합니다.",
            "제공된 학교 자료를 바탕으로 먼저 끝낼 유형과 다시 확인할 유형을 구분합니다.",
        ]
        original = match.group(0)
        leading = " " if original[:1].isspace() else ""
        return leading + options[seed_for(CATEGORY, local, "sanitize-school", original) % len(options)]

    value = re.sub(
        r"[^<>.!?]*(?:(?:학교별|지역\s*학생들의?)[^<>.!?]{0,35}(?:출제\s*(?:흐름|경향)|시험\s*범위)|출제\s*(?:흐름|경향))[^<>.!?]*(?:맞춰|반영|기준)[^<>.!?]*(?:[.!?]|$)",
        replace_school_claim,
        value,
        flags=re.I,
    )

    # 제목·목록처럼 마침표가 없는 짧은 광고형 표현도 남기지 않습니다.
    value = re.sub(
        r"(?:성적|점수)[’']?(?:을|를|으로|로|의)?\s*(?:관리로\s*)?(?:상승|향상|올리(?:는|기|도록|고|며)?|끌어올리(?:는|기|도록|고|며)?|연결(?:되|하)?(?:는|기|도록|고|며)?|만들(?:기|어|도록)?|보장)",
        "학습 과정 점검",
        value,
    )
    value = re.sub(r"성적(?:을)?\s*만드는", "학습 과정을 점검하는", value)
    value = re.sub(
        r"(?:성적|점수)(?:을|를)?\s*위한",
        "학습 개선을 위한",
        value,
    )
    value = re.sub(
        r"(?:학교별\s*)?시험\s*범위[^<>.!?]{0,25}(?:맞춰|반영)",
        "학생이 가져온 시험 범위를 확인한 뒤",
        value,
    )
    value = re.sub(
        r"지역\s*학생[^<>.!?]{0,30}시험\s*범위[^<>.!?]{0,15}맞춰",
        "현재 학교 자료의 시험 범위를 확인한 뒤",
        value,
    )
    value = re.sub(r"출제\s*(?:흐름|경향)(?:을)?\s*(?:반영|맞춰)?", "학교 자료의 유형", value)

    # 드문 비문장형 교차과목 잔여 키워드는 검색 의도에 맞는 중립 표현으로 마무리합니다.
    value = re.sub(r"영어|국어|영수", "수학", value)
    return value


def contextualize_manuscript(
    raw: str,
    row: dict[str, str],
    repeated: set[str],
) -> str:
    local = row_value(row, "근처 수업가능 동네")
    region = row_value(row, "지역")
    district = row_value(row, "시or구")
    center = row_value(row, "센터명") or f"{local} 학습센터"
    schools = split_items(row_value(row, "타깃학교\n(중)"))
    school_reference = schools[seed_for(local, "school") % len(schools)] if schools else "현재 학교 자료"

    raw = sanitize_manuscript(raw.strip(), row)
    raw = re.sub(
        r'^\s*<main class="article-main">',
        '<section class="shell academy-section article-main manuscript-panel">',
        raw,
        count=1,
        flags=re.I,
    )
    raw = re.sub(r"</main>\s*$", "</section>", raw, count=1, flags=re.I)
    inner_heading = f"{region} {district} {local} 중2 수학 학습 설계"
    raw = re.sub(r"<h1>.*?</h1>", f"<h2>{esc(inner_heading)}</h2>", raw, count=1, flags=re.DOTALL | re.I)

    context_templates = [
        "{local}에서는 {evidence}의 최근 풀이 기록과 학교 진도를 함께 놓고 다음 점검 순서를 정하는 것이 좋습니다.",
        "이 기준을 {local} 학생에게 적용할 때는 {evidence} 자료와 오답 원인을 나란히 확인해야 우선순위가 분명해집니다.",
        "{region} {district} {local} 상담에서는 {evidence}에서 확인되는 단원과 주간 복습 가능 시간을 함께 기록해 두는 편이 실용적입니다.",
        "{local} 학부모는 이 항목을 비교할 때 {evidence} 자료를 어떤 방식으로 수업과 재점검에 반영하는지 질문할 수 있습니다.",
        "학생별 차이를 확인하려면 {local}에서 {evidence}의 풀이 흔적과 혼자 다시 풀 수 있는 문제를 구분해 보는 과정이 필요합니다.",
        "{center} 상담에서는 {evidence_object} 기준으로 설명이 필요한 개념과 반복할 문제를 따로 정리할 수 있습니다.",
        "{local}의 실제 학습표에는 {evidence} 점검 결과와 다음 오답 확인 날짜가 함께 적혀야 수업 뒤의 복습까지 이어집니다.",
        "같은 중2 과정이라도 {local} 학생의 학교 진도와 {evidence} 기록에 따라 먼저 보완할 내용은 달라질 수 있습니다.",
        "{district} 지역에서 이 기준을 살필 때는 {evidence} 자료를 근거로 현재 이해와 단순 암기를 구분해 보는 것이 중요합니다.",
        "{local}에서는 설명을 들은 문제와 스스로 해결한 문제를 {evidence}와 함께 비교하면 다음 학습량을 더 구체적으로 정할 수 있습니다.",
    ]
    paragraph_index = 0

    def replace_paragraph(match: re.Match[str]) -> str:
        nonlocal paragraph_index
        attrs, body = match.group(1) or "", match.group(2)
        signature = paragraph_signature(body, local)
        if signature not in repeated or len(signature) < 35:
            paragraph_index += 1
            return match.group(0)
        template = context_templates[seed_for(local, str(paragraph_index), signature) % len(context_templates)]
        context = template.format(
            local=local,
            region=region,
            district=district,
            center=center,
            evidence=school_reference,
            evidence_object=object_form(school_reference),
        )
        paragraph_index += 1
        return f"<p{attrs}>{body.rstrip()} {esc(context)}</p>"

    return re.sub(
        r"<p(\s[^>]*)?>(.*?)</p>",
        replace_paragraph,
        raw,
        flags=re.DOTALL | re.I,
    )


def load_image_rows() -> dict[str, dict[str, str]]:
    return {row.get("제목", "").strip(): row for row in read_csv(IMAGE_CSV)}


def representative_urls() -> list[str]:
    text = REPRESENTATIVE_CSV.read_text(encoding="utf-8-sig")
    urls = re.findall(r'https://[^"\s,>]+\.(?:jpe?g|png|webp|gif)', text, re.I)
    return list(dict.fromkeys(urls))


def find_parent_page(row: dict[str, str]) -> Path | None:
    region = row_value(row, "지역")
    district = row_value(row, "시or구")
    local = row_value(row, "근처 수업가능 동네")
    candidates = [
        SITE / "전국학원" / region / district / local / "index.html",
        SITE / "전국학원" / region / "시" / local / "index.html",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    matches = list((SITE / "전국학원" / region).glob(f"*/{local}/index.html"))
    return matches[0] if matches else None


def parent_relative_url(row: dict[str, str], child: str | None = None) -> str:
    parent = find_parent_page(row)
    if parent:
        relative = parent.parent.relative_to(SITE).as_posix()
    else:
        relative = "/".join(
            [
                "전국학원",
                row_value(row, "지역"),
                row_value(row, "시or구"),
                row_value(row, "근처 수업가능 동네"),
            ]
        )
    if child:
        relative += f"/{child}"
    return "../../../" + "/".join(quote(part, safe="") for part in relative.split("/")) + "/index.html"


def parent_canonical_url(row: dict[str, str], child: str | None = None) -> str:
    parent = find_parent_page(row)
    if parent:
        parts = list(parent.parent.relative_to(SITE).parts)
    else:
        parts = [
            "전국학원",
            row_value(row, "지역"),
            row_value(row, "시or구"),
            row_value(row, "근처 수업가능 동네"),
        ]
    if child:
        parts.append(child)
    return canonical(*parts)


def representative_for(row: dict[str, str], fallback_urls: list[str]) -> str:
    parent = find_parent_page(row)
    candidates: list[Path] = []
    if parent:
        candidates.extend([parent.parent / "중1수학학원" / "index.html", parent])
    for candidate in candidates:
        if not candidate.exists():
            continue
        source = candidate.read_text(encoding="utf-8")
        match = re.search(
            r'data-role="representative-image"[^>]*src="([^"]+)"', source, re.I
        )
        if match:
            return html.unescape(match.group(1))
    local = row_value(row, "근처 수업가능 동네")
    return fallback_urls[seed_for(local, "representative") % len(fallback_urls)]


def map_filename(row: dict[str, str], image_row: dict[str, str]) -> str:
    parent = find_parent_page(row)
    if parent:
        source = parent.read_text(encoding="utf-8")
        match = re.search(r'assets/maps/([^"?#]+)', source, re.I)
        if match and (SITE / "assets" / "maps" / match.group(1)).exists():
            return match.group(1)
    requested = (image_row.get("지도") or "").strip()
    if requested and (SITE / "assets" / "maps" / requested).exists():
        return requested
    slug = row_value(row, "동 영어")
    for suffix in (".jpg", ".jpeg", ".png", ".webp"):
        candidate = SITE / "assets" / "maps" / f"{slug}{suffix}"
        if candidate.exists():
            return candidate.name
    raise FileNotFoundError(f"지도 이미지 없음: {row_value(row, '근처 수업가능 동네')}")


def page_profile(row: dict[str, str]) -> dict[str, str]:
    local = row_value(row, "근처 수업가능 동네")
    profiles = [
        "개념 설명은 이해하지만 문제 조건을 식으로 옮기는 데 시간이 오래 걸리는 학생",
        "숙제는 끝내지만 틀린 문제를 다시 풀지 않아 같은 유형을 반복해서 놓치는 학생",
        "계산은 빠르지만 풀이 단계를 생략해 서술형에서 감점이 생기는 학생",
        "시험 범위가 정해진 뒤에도 복습 순서를 잡지 못해 앞 단원을 놓치는 학생",
        "어려운 문제보다 기본 유형의 작은 실수가 점수를 흔드는 학생",
        "수업 중에는 풀지만 며칠 뒤 혼자 풀 때 풀이가 이어지지 않는 학생",
        "문제량은 충분하지만 풀이 근거를 설명하는 연습이 부족한 학생",
        "현재 진도와 이전 학년의 개념 공백을 함께 점검해야 하는 학생",
        "시험 직전에만 공부해 단원별 오답이 누적되는 학생",
        "풀이 시간이 일정하지 않아 시험 시간 배분이 어려운 학생",
        "응용 문제에서 어떤 개념을 꺼내야 하는지 판단이 느린 학생",
        "학습 계획은 세우지만 완료 여부를 기록하지 않아 복습이 밀리는 학생",
    ]
    priorities = [
        "최근 풀이에서 개념·계산·조건 해석 오류를 먼저 구분합니다.",
        "학교 진도와 누적 오답을 한 표에 놓고 복습 순서를 정합니다.",
        "답만 맞히는 연습보다 풀이 근거를 말하고 쓰는 과정을 확인합니다.",
        "기본 유형의 재풀이 성공 여부를 확인한 뒤 응용 범위를 넓힙니다.",
        "시험일까지 남은 기간을 기준으로 단원별 완료 기준을 나눕니다.",
        "수업 직후·주중·시험 전으로 오답을 다시 볼 시점을 정합니다.",
        "문제별 풀이 시간과 검산 순서를 기록해 시간 배분을 점검합니다.",
        "현재 단원을 설명하는 데 필요한 이전 개념부터 짧게 복원합니다.",
    ]
    checks = [
        "최근 시험지에서 틀린 이유를 학생이 직접 설명할 수 있는지",
        "현재 교재의 진도와 학교 시험 범위가 얼마나 연결되어 있는지",
        "오답을 다시 풀었을 때 해설 없이 해결할 수 있는지",
        "주중에 확보할 수 있는 실제 수학 공부 시간이 어느 정도인지",
        "개념 문제와 응용 문제 중 어느 구간에서 풀이가 멈추는지",
        "풀이 과정에서 부호·조건·계산 순서를 어떻게 검산하는지",
        "시험 전 복습 계획이 단원별 완료 기준으로 적혀 있는지",
        "질문이 생겼을 때 표시하고 다음 수업에서 확인하는 습관이 있는지",
    ]
    return {
        "student": choose(local, "student", profiles),
        "priority": choose(local, "priority", priorities),
        "check": choose(local, "check", checks),
    }


def build_subject_peer_network(
    rows: list[dict[str, str]],
    degree: int = 6,
) -> dict[str, list[str]]:
    """시군구로 정렬한 원형 그래프에서 정확히 6개의 상호 형제 링크를 만듭니다."""
    if degree <= 0 or degree % 2:
        raise ValueError("상호 링크 차수는 0보다 큰 짝수여야 합니다.")
    if len(rows) <= degree:
        raise ValueError("지역 수가 상호 링크 차수보다 많아야 합니다.")

    indexed_rows = list(enumerate(rows))
    indexed_rows.sort(
        key=lambda pair: (
            row_value(pair[1], "지역"),
            row_value(pair[1], "시or구"),
            pair[0],
        )
    )
    order = [row_value(row, "근처 수업가능 동네") for _index, row in indexed_rows]
    graph: dict[str, set[str]] = {local: set() for local in order}
    half = degree // 2
    for index, local in enumerate(order):
        for distance in range(1, half + 1):
            graph[local].add(order[(index - distance) % len(order)])
            graph[local].add(order[(index + distance) % len(order)])

    if any(len(peers) != degree for peers in graph.values()):
        raise ValueError("과목별 지역 링크망에서 정확히 6개의 상호 링크를 확보하지 못했습니다.")
    if any(local not in graph[peer] for local, peers in graph.items() for peer in peers):
        raise ValueError("과목별 지역 링크망이 상호 왕복 구조가 아닙니다.")
    return {local: sorted(peers, key=order.index) for local, peers in graph.items()}


def build_math_enrichment(row: dict[str, str], profile: dict[str, str]) -> str:
    """원고를 대체하지 않고 실제 센터 자료와 학생 상황을 연결해 깊이를 보강합니다."""
    local = row_value(row, "근처 수업가능 동네")
    region = row_value(row, "지역")
    district = row_value(row, "시or구")
    center = row_value(row, "센터명") or f"{local} 학습센터"
    schools = split_items(row_value(row, "타깃학교\n(중)"))
    evidence = schools[seed_for(CATEGORY, local, "enrichment-school") % len(schools)] if schools else "현재 학교의 시험지와 교재"

    diagnostic_signals = [
        "정답은 맞지만 풀이의 첫 단계를 설명하지 못하는 문제",
        "부호와 괄호를 옮기는 과정에서 반복되는 계산 실수",
        "문제 조건을 읽고도 사용할 개념을 고르는 데 오래 걸리는 유형",
        "수업 직후에는 풀지만 며칠 뒤 풀이 순서를 떠올리지 못하는 문항",
        "도형의 조건을 그림에 표시하지 않아 놓치는 관계",
        "서술형에서 식은 세웠지만 근거 문장을 빠뜨리는 답안",
        "이전 단원의 개념이 현재 단원 풀이를 막는 연결 지점",
        "시험 후 오답을 답만 고치고 원인을 기록하지 않은 문항",
        "풀이 시간은 길지만 검산 순서가 정해져 있지 않은 문제",
        "기본 유형과 변형 유형의 공통 조건을 구분하지 못하는 문항",
        "학습량은 충분하지만 완료 여부가 기록되지 않은 과제",
        "질문 표시 없이 해설을 먼저 읽어 혼자 해결한 범위를 알기 어려운 문제",
    ]
    evidence_actions = [
        "풀이를 개념 이해·조건 해석·계산·검산 네 단계로 나누어 멈춘 위치를 표시합니다.",
        "같은 단원에서 맞힌 문제와 틀린 문제를 함께 놓고 풀이 과정의 차이를 학생이 설명하게 합니다.",
        "해설을 보기 전 시도한 식과 질문 표시를 남겨 실제로 필요한 설명 범위를 구분합니다.",
        "오답 원인을 한 단어로 적지 않고 잘못 읽은 조건과 다시 사용할 개념을 각각 기록합니다.",
        "현재 문제를 푸는 데 필요한 이전 개념만 짧게 복원한 뒤 바로 유사 문항에 적용합니다.",
        "풀이 시간을 재기 전에 정확한 순서를 재현할 수 있는지 확인하고 이후 시간 기준을 세웁니다.",
        "학생이 말로 설명한 풀이와 종이에 쓴 식이 같은 흐름인지 대조합니다.",
        "교재 진도와 학교 범위를 한 표에 적어 먼저 끝낼 항목과 다시 볼 항목을 분리합니다.",
        "틀린 문제를 난도별로 묶지 않고 오류 원인별로 묶어 다음 복습 순서를 정합니다.",
        "답을 가린 재풀이와 숫자·조건을 바꾼 적용 문제를 구분해 이해 여부를 확인합니다.",
        "과제 완료 시각과 막힌 문제를 함께 기록해 다음 수업의 설명 시간을 조정합니다.",
        "검산할 항목을 부호·조건·단위·답의 범위로 나누어 시험지에 적용합니다.",
    ]
    verification_actions = [
        "이틀 뒤 해설 없이 같은 풀이를 다시 구성할 수 있는지 확인합니다.",
        "숫자나 조건이 달라진 문제에서도 같은 개념을 선택하는지 살펴봅니다.",
        "정답보다 풀이 근거를 먼저 말하게 해 우연히 맞힌 문제를 구분합니다.",
        "다음 과제에서 같은 오류가 줄었는지 오답 기록과 나란히 비교합니다.",
        "제한 시간을 적용하기 전 정확도를 유지하는 문제 수를 확인합니다.",
        "학생이 스스로 질문을 만들고 필요한 개념을 찾아 설명하는지 봅니다.",
        "시험 범위 안에서 다시 등장한 유형에 같은 검산 기준을 적용하는지 확인합니다.",
        "한 주 뒤 누적 확인 문제에서 풀이 순서를 재현하는지 점검합니다.",
        "틀렸던 이유와 고친 이유를 서로 다른 문장으로 설명할 수 있는지 확인합니다.",
        "교재 표시 없이도 단원의 핵심 조건을 찾아 식으로 옮기는지 살펴봅니다.",
        "완료 기록과 실제 재풀이 결과가 일치하는지 다음 수업에서 대조합니다.",
        "학교 자료의 새 문항에서 배운 기준을 독립적으로 적용하는지 확인합니다.",
    ]
    unit_focuses = [
        ("식과 계산", "정답만 빠르게 내기보다 부호·괄호·계산 순서를 줄마다 확인하고 틀린 지점을 특정합니다."),
        ("방정식과 함수", "조건을 식으로 바꾸는 과정과 표·식·그래프 사이의 연결을 말로 설명하게 합니다."),
        ("도형", "주어진 조건을 그림에 표시하고 사용할 성질과 결론 사이의 근거를 순서대로 적습니다."),
        ("확률과 자료", "경우를 빠뜨리지 않는 분류 기준과 표·그래프에서 읽은 수치의 의미를 함께 확인합니다."),
        ("서술형", "계산식뿐 아니라 사용한 개념, 조건과 결론이 답안에 드러나는지 단계별로 점검합니다."),
        ("누적 복습", "현재 단원을 막는 이전 개념만 골라 복원하고 학교 진도와 별도 일정으로 재확인합니다."),
    ]
    unit_focuses.sort(key=lambda item: seed_for(CATEGORY, local, "unit-focus", item[0]))

    signal = choose(local, "diagnostic-signal", diagnostic_signals)
    action = choose(local, "evidence-action", evidence_actions)
    verification = choose(local, "verification-action", verification_actions)
    diagnosis_intro = choose(
        local,
        "diagnosis-intro",
        [
            f"{local}에서 중2 수학의 시작점을 정할 때는 단원명보다 최근 풀이에서 반복된 행동을 보는 편이 정확합니다.",
            f"{region} {district}의 같은 중2 과정이라도 학생마다 막히는 단계가 다르므로 최근 답안과 재풀이 결과를 함께 확인해야 합니다.",
            f"{center} 상담에서 학습량을 정하기 전, 학생이 혼자 해결한 범위와 설명을 들은 뒤 해결한 범위를 구분해 두는 것이 좋습니다.",
            f"{evidence} 자료를 기준으로 보면 현재 진도와 누적 공백 중 무엇을 먼저 다뤄야 하는지 더 구체적으로 정할 수 있습니다.",
            f"중2 수학은 문제 수보다 오류가 시작된 위치를 찾는 일이 우선입니다. {local} 상담에서도 풀이 흔적을 근거로 순서를 정해야 합니다.",
            f"시험 범위가 같더라도 복습 시간과 오답 유형은 다릅니다. {local} 학생의 실제 기록을 바탕으로 완료 기준을 나누는 과정이 필요합니다.",
        ],
    )
    plan_sets = [
        [("1주차", "최근 시험지와 교재에서 오류 원인을 분류합니다."), ("2주차", "우선 개념을 짧게 복원하고 기본 유형에 적용합니다."), ("3주차", "학교 진도 문항과 변형 문제에서 풀이 근거를 확인합니다."), ("4주차", "누적 오답을 해설 없이 다시 풀고 다음 계획을 조정합니다.")],
        [("진단 주간", "맞힌 문제까지 포함해 풀이 설명이 가능한 범위를 구분합니다."), ("연결 주간", "이전 개념과 현재 단원의 연결 문제를 집중해서 다룹니다."), ("적용 주간", "조건이 달라진 문항과 서술형에서 같은 기준을 적용합니다."), ("재확인 주간", "간격을 둔 재풀이 결과와 학습 기록을 대조합니다.")],
        [("첫 확인", "학교 범위와 학생의 누적 오답을 한 표에 정리합니다."), ("핵심 보완", "계산·조건 해석·개념 중 우선 오류를 집중해서 보완합니다."), ("독립 풀이", "질문과 해설 없이 해결할 수 있는 문제 범위를 넓힙니다."), ("시험 연결", "검산 순서와 시간 배분을 학교 자료에 적용합니다.")],
        [("자료 정리", "시험지·교재·과제의 틀린 문제를 원인별로 묶습니다."), ("개념 복원", "현재 진도에 꼭 필요한 공백만 골라 다시 설명합니다."), ("유형 확장", "기본 풀이가 안정된 뒤 조건을 바꾼 문제로 확장합니다."), ("완료 검증", "재풀이 날짜와 성공 여부를 기록해 다음 분량을 정합니다.")],
        [("기준 세우기", "학생이 설명할 수 있는 풀이와 외운 풀이를 나눕니다."), ("과정 고치기", "틀린 지점부터 풀이 순서와 검산 방식을 다시 설계합니다."), ("학교 자료 적용", "현재 진도와 시험 범위 안에서 동일 기준을 반복합니다."), ("누적 확인", "한 주 뒤 다시 풀어 기억이 아닌 이해 여부를 확인합니다.")],
        [("범위 나누기", "남은 기간과 단원별 완료 기준을 현실적인 분량으로 정합니다."), ("약점 우선", "가장 자주 반복된 오류를 먼저 줄이는 연습을 진행합니다."), ("실전 연결", "서술형과 시간 제한 문제에 검산 순서를 적용합니다."), ("기록 갱신", "새로운 오답과 해결된 오답을 나눠 계획표를 갱신합니다.")],
    ]
    plan = choose(local, "four-step-plan", plan_sets)
    plan_html = "".join(
        f'<article class="article-target-card"><h3>{esc(label)}</h3><p>{esc(body)} {esc(choose(local, "plan-" + label, verification_actions))}</p></article>'
        for label, body in plan
    )
    unit_html = "".join(
        f'<article class="article-card"><h3>{esc(title)}</h3><p>{esc(body)} {esc(choose(local, "unit-" + title, evidence_actions))}</p></article>'
        for title, body in unit_focuses[:3]
    )

    return f'''<section class="shell academy-section article-main subject-enrichment-panel">
      <section class="article-section article-local-feature-section">
        <p class="article-eyebrow">INDIVIDUAL DIAGNOSIS</p>
        <h2>{esc(local)} 중2 수학 진단을 실제 기록으로 좁히는 방법</h2>
        <p>{esc(diagnosis_intro)} 특히 {esc(signal)}를 찾아야 합니다. {esc(action)} {esc(verification)}</p>
        <div class="article-card-grid">{unit_html}</div>
      </section>
      <section class="article-section article-local-feature-section">
        <h2>{esc(object_form(evidence))} 기준으로 살펴볼 4단계 실행 흐름</h2>
        <p>{esc(profile['student'])}에게 적용할 수 있는 비교용 예시입니다. 실제 진도와 분량은 진단 결과, 학교 일정과 주중 복습 가능 시간을 확인한 뒤 조정해야 합니다.</p>
        <div class="article-target-list">{plan_html}</div>
      </section>
      <section class="article-closing"><p>{esc(local)} 학부모가 수업을 비교할 때는 진도표만 보지 말고 {esc(profile['check'])}, 재풀이 날짜와 완료 기록이 다음 수업에 실제로 반영되는지 확인하는 것이 좋습니다.</p></section>
    </section>'''


def build_faqs(row: dict[str, str], profile: dict[str, str]) -> list[tuple[str, str]]:
    local = row_value(row, "근처 수업가능 동네")
    title = f"{local} 중2 수학학원"
    center = row_value(row, "센터명") or f"{local} 학습센터"
    schools = split_items(row_value(row, "타깃학교\n(중)"))
    grades = split_items(row_value(row, "가능학년\n(수학)"))
    available = "중2" in grades
    school_answer = (
            f"공개 센터 자료에는 {', '.join(schools)} 등이 수업 가능 학교로 안내되어 있습니다. "
        "학교별 시험 범위와 일정은 달라질 수 있으므로 상담할 때 현재 학교 자료를 함께 확인하는 것이 좋습니다."
        if schools
        else "공개 센터 자료에는 중학교명이 별도로 기재되어 있지 않습니다. 학교명을 임의로 단정하지 않고 상담 시 현재 학교와 시험 범위를 확인합니다."
    )
    availability_answer = (
        f"공개된 {center} 가능 학년 정보에 중2 수학이 포함되어 있습니다. 실제 시간표, 반 편성 및 시작 가능일은 변동될 수 있어 상담 시 다시 확인해야 합니다."
        if available
        else f"공개된 {center} 자료에는 수학 가능 학년이 기재되어 있지 않습니다. 중2 수학 개설 여부와 수업 일정은 상담을 통해 확인해야 합니다."
    )
    bank = [
        (
            f"{title} 상담 전에 무엇을 준비하면 좋나요?",
            f"최근 시험지, 현재 수학 교재, 학교 진도와 반복 오답을 준비하면 좋습니다. 특히 {profile['check']}를 메모해 가면 상담 기준을 구체적으로 세울 수 있습니다.",
        ),
        (
            f"{local} 중2 수학학원은 어떤 학생에게 맞는지 어떻게 판단하나요?",
            f"{profile['student']}이라면 진단 과정과 재풀이 확인 방식을 먼저 살펴보세요. 학원 이름이나 문제량만 비교하기보다 학생이 혼자 풀 수 있게 되는 과정을 확인해야 합니다.",
        ),
        (
            "중2 수학 내신은 진도와 복습 중 무엇을 먼저 봐야 하나요?",
            f"현재 학교 진도만 앞서가기보다 이전 단원의 공백과 최근 오답을 먼저 구분해야 합니다. {profile['priority']}",
        ),
        (
            "오답 관리는 어떤 방식인지 확인해야 하나요?",
            "틀린 문제를 다시 푸는 것에 그치지 않고 개념 부족, 계산 실수, 조건 해석, 풀이 순서 중 원인을 구분하는지 확인하세요. 이후 해설 없이 재풀이하는 날짜가 정해지는지도 중요합니다.",
        ),
        (
            "학교별 내신 대비 자료도 확인할 수 있나요?",
            school_answer,
        ),
        (
            "중2 수학 수강 가능 여부는 어디에서 확인하나요?",
            availability_answer,
        ),
        (
            "선행보다 복습이 필요한 학생도 상담할 수 있나요?",
            "현재 학년 진도를 무조건 앞당기기보다 문제를 푸는 데 필요한 이전 개념을 찾아 현재 단원과 연결하는 방향으로 상담할 수 있습니다. 실제 적용 범위는 진단 결과를 바탕으로 확인합니다.",
        ),
        (
            "상담할 때 수업 시간표와 교습비도 확인해야 하나요?",
            "네. 공개된 교습비 자료와 함께 주당 횟수, 시작·종료 시각, 결석·보강 기준, 시험 기간 일정 변동을 같은 표에 적어 비교하는 것이 좋습니다.",
        ),
        (
            "중2 수학에서 계산 실수와 개념 부족은 어떻게 구분하나요?",
            "같은 유형을 다시 풀 때 식을 세우는 단계부터 막히면 개념과 조건 해석을, 풀이 흐름은 맞지만 부호·괄호에서 틀리면 계산과 검산 습관을 우선 확인할 수 있습니다.",
        ),
        (
            "시험 범위가 나오기 전에는 어떤 공부를 해야 하나요?",
            "현재 진도를 따라가면서 최근 오답과 이전 단원의 공백을 정리하는 것이 좋습니다. 범위가 확정되면 교과서·학교 자료와 서술형, 시간 배분 중심으로 계획을 전환합니다.",
        ),
        (
            "숙제의 양보다 먼저 확인할 기준이 있나요?",
            "학생이 실제로 끝낼 수 있는 분량인지, 막힌 문제를 표시하는지, 제출 뒤 오답을 다시 풀 날짜가 정해지는지를 확인해야 합니다. 완료 기준이 문제 수에만 머물지 않는 것이 중요합니다.",
        ),
        (
            "중2 수학 학습 계획은 얼마나 자주 조정해야 하나요?",
            "고정된 계획을 오래 유지하기보다 주간 재풀이 결과와 학교 진도를 확인해 해결된 항목과 남은 항목을 나누고 다음 분량을 조정하는 방식이 실용적입니다.",
        ),
    ]
    ordered = sorted(bank, key=lambda item: seed_for(CATEGORY, local, "faq", item[0]))
    required = [bank[0], bank[1], bank[4], bank[5]]
    selected = required + [item for item in ordered if item not in required][:3]
    selected.sort(key=lambda item: seed_for(CATEGORY, local, "faq-order", item[0]))
    context_notes = [
        f"{local}에서는 최근 풀이 기록과 실제 주중 일정을 함께 놓고 이 기준을 확인하세요.",
        f"{center} 상담에서는 이 항목이 다음 과제와 재풀이 날짜에 어떻게 반영되는지도 질문하는 것이 좋습니다.",
        f"{profile['student']}이라면 문제량보다 혼자 다시 해결한 범위를 근거로 판단해야 합니다.",
        f"상담 전에 {profile['check']}를 적어 두면 {local} 수업 방식을 더 구체적으로 비교할 수 있습니다.",
        f"이 기준은 {local} 학생의 학교 진도와 복습 가능 시간에 맞춰 조정해야 합니다.",
        f"실제 적용 여부는 {center}의 반 편성, 시간표와 학생 진단 결과를 확인한 뒤 결정합니다.",
        f"{profile['priority'].rstrip('.!? ')}라는 원칙이 수업 기록에 남는지도 함께 살펴보세요.",
        f"학교 자료가 있다면 답만 보지 말고 풀이 흔적과 다시 푼 결과를 나란히 확인하는 편이 좋습니다.",
    ]
    ordered_context_notes = sorted(
        context_notes,
        key=lambda note: seed_for(CATEGORY, local, "faq-context-order", note),
    )
    return [
        (
            question,
            f"{answer} {ordered_context_notes[index]}",
        )
        for index, (question, answer) in enumerate(selected)
    ]


def build_parent_notes(row: dict[str, str], profile: dict[str, str]) -> list[str]:
    local = row_value(row, "근처 수업가능 동네")
    center = row_value(row, "센터명") or f"{local} 학습센터"
    schools = split_items(row_value(row, "타깃학교\n(중)"))
    grades = split_items(row_value(row, "가능학년\n(수학)"))
    tuition = row_value(row, "센터 교습비")
    school_reference = ", ".join(schools[:2]) if schools else "현재 학교의 시험지와 교재"
    grade_note = "공개 자료에 중2 수학이 기재되어 있는지" if "중2" in grades else "중2 수학 개설 여부를"
    tuition_note = "공개 교습비 링크와" if tuition else "상담에서 교습비와"
    options = [
        f"{local} 상담에서는 문제 수보다 학생이 틀린 이유를 설명하고 해설 없이 다시 풀 수 있는지 확인하는 편이 적절합니다.",
        f"{profile['student']}이라면 진도를 서두르기보다 주간 복습 시점과 오답 확인 날짜가 실제 기록에 남는지 살펴볼 수 있습니다.",
        f"{center} 공개 정보를 비교할 때는 학교 진도, 수학 가능 학년과 교습비 확인 경로를 같은 표에 정리하는 것이 좋습니다.",
        f"{local} 중2 수학학원은 수업 시간뿐 아니라 과제의 완료 기준과 재풀이 결과를 다음 수업에서 확인하는지 함께 비교해야 합니다.",
        f"{object_form(school_reference)} 준비하면 개념 부족, 조건 해석과 계산 실수를 실제 답안으로 구분하는 데 도움이 됩니다.",
        "시험 범위가 정해지기 전에는 누적 공백을, 범위가 확정된 뒤에는 학교 자료와 오답의 완료 순서를 확인하는 것이 좋습니다.",
        "수업에서 맞힌 문제도 며칠 뒤 혼자 다시 풀 수 있는지 확인해야 설명을 기억한 문제와 이해한 문제를 구분할 수 있습니다.",
        f"{local}에서는 통학 여부와 함께 종료 시각, 결석·보강 기준과 시험 기간 시간표를 실제 주간 일정에 맞춰 살펴봐야 합니다.",
        f"{profile['check']}를 상담 질문으로 적어 두면 문제량보다 관리 방식의 차이를 구체적으로 비교할 수 있습니다.",
        f"{profile['priority'].rstrip('.!? ')}라는 기준이 주간 계획과 다음 과제에 어떻게 반영되는지 확인하는 것이 중요합니다.",
        f"{grade_note} 확인하고, 실제 반 편성·시작 가능일은 {center} 상담에서 다시 확인해야 합니다.",
        f"{tuition_note} 주당 횟수, 수업 시간, 결석·보강 조건을 함께 확인해야 실제 운영 조건을 비교할 수 있습니다.",
        "오답노트의 양보다 오류 원인, 다시 푼 날짜와 성공 여부가 함께 기록되는지 보는 편이 학습 과정을 판단하기 쉽습니다.",
        "시간 제한 문제는 정확한 풀이 순서가 자리 잡은 뒤 적용하는지, 검산 시간이 따로 확보되는지 확인할 필요가 있습니다.",
        f"{local}의 현재 학교 자료가 공개 목록에 없더라도 학교명을 임의로 단정하지 않고 상담에서 실제 자료를 확인해야 합니다.",
        f"{center} 안내만으로 확정할 수 없는 시간표와 반 구성은 학생 진단 결과와 함께 상담에서 확인하는 것이 안전합니다.",
    ]
    options.sort(key=lambda value: seed_for(CATEGORY, local, "parent-view", value))
    return options[:3]


def build_checklist(row: dict[str, str], profile: dict[str, str]) -> list[tuple[str, str]]:
    local = row_value(row, "근처 수업가능 동네")
    center = row_value(row, "센터명") or f"{local} 학습센터"
    schools = split_items(row_value(row, "타깃학교\n(중)"))
    grades = split_items(row_value(row, "가능학년\n(수학)"))
    tuition = row_value(row, "센터 교습비")
    school_reference = schools[seed_for(CATEGORY, local, "checklist-school") % len(schools)] if schools else "현재 학교 자료"
    materials = [
        f"{school_reference}의 최근 시험지와 현재 교재에서 틀린 문제를 표시합니다.",
        f"{local} 상담 전에 최근 시험지, 교재 진도와 학교에서 받은 범위를 한곳에 모읍니다.",
        f"{school_reference} 자료 중 학생이 혼자 푼 문제와 설명을 들은 문제를 구분합니다.",
        "최근 시험지, 오답노트와 현재 사용하는 교재의 진도표를 함께 준비합니다.",
        "시험 범위가 아직 없다면 현재 교재와 이전 시험의 누적 오답부터 정리합니다.",
        "교과서·학교 프린트·과제에서 반복해서 막힌 문항에 질문 표시를 남깁니다.",
        "맞힌 문제까지 포함해 풀이 근거를 설명하기 어려운 문항을 따로 표시합니다.",
        f"{center} 상담에서 확인할 수 있도록 학교 진도와 집에서 복습한 범위를 기록합니다.",
    ]
    errors = [
        f"{profile['check']}를 최근 답안에서 확인합니다.",
        "개념 부족·조건 해석·계산·검산 중 오류가 시작된 위치를 적습니다.",
        "해설을 보기 전 남긴 식과 질문 표시를 기준으로 막힌 단계를 구분합니다.",
        "같은 유형을 다시 풀어 식 세우기와 계산 중 어느 단계가 달라지는지 봅니다.",
        "틀린 이유와 고친 이유를 서로 다른 문장으로 설명할 수 있는지 확인합니다.",
        "부호·괄호·조건 누락과 풀이 순서 오류를 한 항목으로 묶지 않습니다.",
        "기본 문제와 조건이 바뀐 문제의 풀이 차이를 학생이 말할 수 있는지 봅니다.",
        f"{profile['student']}에게 먼저 확인할 오류를 한 가지로 좁혀 기록합니다.",
    ]
    schedules = [
        "수업 직후, 주중과 시험 전으로 재풀이 날짜를 나누어 적습니다.",
        "문제 수보다 실제로 복습할 수 있는 요일과 시간을 먼저 계산합니다.",
        "한 주 뒤 누적 확인 문제를 풀 시간을 주간 계획에 따로 확보합니다.",
        "학교 일정과 과제량을 놓고 기본·적용·서술형 완료 시점을 나눕니다.",
        "시험 범위 확정 전과 확정 후의 복습 시간을 서로 다른 계획으로 적습니다.",
        "막힌 문제를 질문할 시간과 답을 가리고 다시 풀 시간을 각각 정합니다.",
        "정확도가 안정된 뒤 시간 제한과 검산 시간을 적용할 날짜를 정합니다.",
        f"{local}의 실제 통학·귀가 시간을 제외하고 확보 가능한 복습 시간을 계산합니다.",
    ]
    operations = [
        f"{center}의 반 편성, 시작 가능일과 결석·보강 기준을 상담에서 확인합니다.",
        f"{'공개 교습비 링크' if tuition else '상담 안내'}와 주당 횟수·수업 시간을 같은 표에서 비교합니다.",
        f"{'중2 수학 기재 내용을' if '중2' in grades else '중2 수학 개설 여부를'} 확인하고 실제 시간표는 상담에서 다시 확인합니다.",
        "시험 기간의 일정 변경, 보강 방식과 과제 피드백 시점을 질문합니다.",
        "반 구성만 보지 않고 질문 확인 방식과 재풀이 결과의 기록 여부를 살펴봅니다.",
        "수업 시작·종료 시각과 귀가 시간을 주간 학습 계획에 함께 적습니다.",
        "교습비, 수업 횟수와 결석 시 보강 조건을 각각 확인해 비교합니다.",
        f"공개 자료로 확정할 수 없는 운영 조건은 {center} 상담에서 확인합니다.",
    ]
    material_body = choose(local, "checklist-materials", materials)
    material_context = choose(
        local,
        "checklist-material-context",
        [
            f"{local} 상담에서는 이 자료의 풀이 흔적을 기준으로 확인합니다.",
            f"준비한 자료는 {center}에서 실제 진도와 대조할 때 활용할 수 있습니다.",
            f"{local} 학생의 최근 답안과 재풀이 결과를 나란히 놓고 살펴봅니다.",
            f"자료에 없는 학교 일정은 임의로 단정하지 않고 {local} 상담에서 확인합니다.",
        ],
    )
    return [
        ("준비 자료", f"{material_body} {material_context}"),
        ("오답 원인", choose(local, "checklist-errors", errors)),
        ("주간 일정", choose(local, "checklist-schedule", schedules)),
        ("운영 조건", choose(local, "checklist-operations", operations)),
    ]


def page_json_ld(
    row: dict[str, str],
    description: str,
    page_url: str,
    rep_image: str,
    body_image: str,
    map_image: str,
    faqs: list[tuple[str, str]],
    related: list[tuple[str, str]],
) -> dict:
    local = row_value(row, "근처 수업가능 동네")
    region = row_value(row, "지역")
    district = row_value(row, "시or구")
    title = f"{local} 중2 수학학원"
    center = row_value(row, "센터명") or f"{local} 학습센터"
    address = row_value(row, "센터 주소")
    registration = row_value(row, "교육지원청 등록번호")
    schools = split_items(row_value(row, "타깃학교\n(중)"))
    grades = split_items(row_value(row, "가능학년\n(수학)"))
    tuition_url = row_value(row, "센터 교습비")

    organization = {
        "@type": ["EducationalOrganization", "LocalBusiness"],
        "@id": f"{page_url}#organization",
        "name": center,
        "url": page_url,
        "telephone": PHONE_DISPLAY,
        "areaServed": {"@type": "Place", "name": f"{region} {district} {local}"},
        "address": {"@type": "PostalAddress", "streetAddress": address, "addressCountry": "KR"},
        "educationalLevel": grades,
        "knowsAbout": ["중2 수학", "중학교 내신", "개념 진단", "오답 재학습", "학습 플래너"],
        "contactPoint": {
            "@type": "ContactPoint",
            "telephone": f"+82-{PHONE_DISPLAY[1:]}",
            "contactType": "교육 상담",
            "availableLanguage": "Korean",
            "url": FORM_URL,
        },
        "makesOffer": [
            {
                "@type": "Offer",
                "name": f"{title} 학습 상담",
                "category": "중2 수학 학습 진단",
                "url": tuition_url or page_url,
                "itemOffered": {"@id": f"{page_url}#service"},
            }
        ],
    }
    if registration:
        organization["identifier"] = registration
    organization["mentions"] = [
        {"@type": "School", "name": school} for school in schools
    ]

    return {
        "@context": "https://schema.org",
        "@graph": [
            {
                "@type": "WebSite",
                "@id": f"{DOMAIN}/#website",
                "url": f"{DOMAIN}/",
                "name": SITE_NAME,
                "inLanguage": "ko-KR",
            },
            organization,
            {
                "@type": "WebPage",
                "@id": f"{page_url}#webpage",
                "url": page_url,
                "name": title,
                "description": description,
                "inLanguage": "ko-KR",
                "isPartOf": {"@id": f"{DOMAIN}/#website"},
                "breadcrumb": {"@id": f"{page_url}#breadcrumb"},
                "mainEntity": {"@id": f"{page_url}#service"},
                "primaryImageOfPage": {"@id": f"{page_url}#primaryimage"},
                "about": [
                    {"@type": "Place", "name": f"{region} {district} {local}"},
                    {"@type": "Thing", "name": "중2 수학학원"},
                    {"@type": "Thing", "name": "중학교 2학년 수학 내신"},
                ],
                "mentions": [
                    {"@type": "EducationalOrganization", "name": center},
                    *[{"@type": "School", "name": school} for school in schools],
                ],
                "hasPart": [
                    {"@type": "WebPageElement", "name": name}
                    for name in ["핵심 답변", "중2 수학 학습 설계", "센터 정보", "상담 전 체크리스트", "FAQ", "학부모 상담 관점", "내부링크"]
                ],
            },
            {
                "@type": "ImageObject",
                "@id": f"{page_url}#primaryimage",
                "url": rep_image,
                "caption": f"{title} {SITE_NAME} 대표",
            },
            {
                "@type": "BreadcrumbList",
                "@id": f"{page_url}#breadcrumb",
                "itemListElement": [
                    {"@type": "ListItem", "position": 1, "name": "홈", "item": f"{DOMAIN}/"},
                    {"@type": "ListItem", "position": 2, "name": PARENT, "item": canonical(PARENT)},
                    {"@type": "ListItem", "position": 3, "name": CATEGORY_LABEL, "item": canonical(PARENT, CATEGORY)},
                    {"@type": "ListItem", "position": 4, "name": title, "item": page_url},
                ],
            },
            {
                "@type": "Service",
                "@id": f"{page_url}#service",
                "name": f"{title} 학습 상담 및 안내",
                "serviceType": "중학교 2학년 수학 학습관리",
                "provider": {"@id": f"{page_url}#organization"},
                "areaServed": {"@type": "Place", "name": f"{region} {district} {local}"},
                "audience": {"@type": "EducationalAudience", "educationalRole": "중학교 2학년 학생 및 학부모"},
                "about": ["중2 수학", "내신 대비", "오답 관리", "학습 습관"],
            },
            {
                "@type": "Article",
                "@id": f"{page_url}#article",
                "url": page_url,
                "headline": title,
                "description": description,
                "datePublished": PUBLISH_DATE,
                "dateModified": PUBLISH_DATE,
                "inLanguage": "ko-KR",
                "mainEntityOfPage": {"@id": f"{page_url}#webpage"},
                "author": {"@id": f"{page_url}#organization"},
                "publisher": {"@id": f"{page_url}#organization"},
                "image": [rep_image, body_image, map_image],
                "articleSection": [region, district, local, "중2 수학", "내신 대비", "오답 재학습"],
                "about": [{"@type": "Thing", "name": title}],
                "mentions": [{"@type": "EducationalOrganization", "name": center}],
            },
            {
                "@type": "FAQPage",
                "@id": f"{page_url}#faq",
                "mainEntity": [
                    {
                        "@type": "Question",
                        "name": question,
                        "acceptedAnswer": {"@type": "Answer", "text": answer},
                    }
                    for question, answer in faqs
                ],
            },
            {
                "@type": "ItemList",
                "@id": f"{page_url}#related-pages",
                "name": f"{local} 관련 학습 페이지",
                "itemListElement": [
                    {"@type": "ListItem", "position": index, "name": name, "url": url}
                    for index, (name, url) in enumerate(related, 1)
                ],
            },
        ],
    }


def head_html(title: str, description: str, page_url: str, image: str, schema: dict, asset_prefix: str, og_type: str = "article") -> str:
    return f'''<!doctype html>
<html lang="ko">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{esc(title)} | {SITE_NAME}</title>
  <meta name="description" content="{esc(description)}">
  <meta name="robots" content="index, follow, max-image-preview:large">
  <link rel="canonical" href="{page_url}">
  <meta property="og:type" content="{esc(og_type)}">
  <meta property="og:site_name" content="{SITE_NAME}">
  <meta property="og:title" content="{esc(title)} | {SITE_NAME}">
  <meta property="og:description" content="{esc(description)}">
  <meta property="og:url" content="{page_url}">
  <meta property="og:image" content="{esc(image)}">
  <meta name="twitter:card" content="summary_large_image">
  <link rel="icon" href="{asset_prefix}assets/favicon.png" type="image/png">
  <link rel="stylesheet" href="{asset_prefix}assets/site.css">
  <script type="application/ld+json">{json.dumps(schema, ensure_ascii=False, separators=(",", ":"))}</script>
</head>'''


def nav_html(prefix: str, active: str) -> str:
    items = [
        ("홈", f"{prefix}index.html", "home"),
        ("학습가이드", f"{prefix}학습가이드/index.html", "guide"),
        ("상담문의", f"{prefix}상담문의/index.html", "contact"),
        ("과목별학원", f"{prefix}과목별학원/index.html", "subject"),
        ("전국학원", f"{prefix}전국학원/index.html", "nationwide"),
    ]
    links = "".join(
        f'<a{" class=\"active\"" if key == active else ""} href="{href}">{label}</a>'
        for label, href, key in items
    )
    return f'''  <a class="skip-link" href="#main">본문 바로가기</a>
  <header class="site-header">
    <a class="brand" href="{prefix}index.html" aria-label="와와센터 홈"><span class="brand-mark">W</span><span><strong>와와센터</strong><small>Learning Control Studio</small></span></a>
    <nav class="top-nav" aria-label="상단 메뉴">{links}</nav>
  </header>'''


def footer_html(asset_prefix: str) -> str:
    return f'''  <aside class="floating-actions" aria-label="빠른 상담 버튼"><a href="tel:{PHONE_LINK}">전화문의</a><a href="{SMS_URL}" target="_blank" rel="noopener noreferrer">문자문의</a><a href="{FORM_URL}" target="_blank" rel="noopener noreferrer">상담신청</a></aside>
  <footer class="site-footer"><div><strong>{SITE_NAME}</strong><p>영어·수학·국어 학습코칭 안내</p></div><a href="tel:{PHONE_LINK}">{PHONE_DISPLAY}</a></footer>
  <script src="{asset_prefix}assets/site.js"></script>'''


def detail_html(
    row: dict[str, str],
    manuscript: str,
    image_row: dict[str, str],
    rep_image: str,
    map_name: str,
    index: int,
    rows: list[dict[str, str]],
    peer_locals: list[str],
) -> str:
    local = row_value(row, "근처 수업가능 동네")
    region = row_value(row, "지역")
    district = row_value(row, "시or구")
    center = row_value(row, "센터명") or f"{local} 학습센터"
    address = row_value(row, "센터 주소")
    location_guide = inline_text(row_value(row, "위치안내"))
    registration_name = row_value(row, "교육지원청명칭")
    registration_number = row_value(row, "교육지원청 등록번호")
    tuition = row_value(row, "센터 교습비")
    grades = split_items(row_value(row, "가능학년\n(수학)"))
    schools = split_items(row_value(row, "타깃학교\n(중)"))
    available = "중2" in grades
    title = f"{local} 중2 수학학원"
    profile = page_profile(row)
    description = (
        f"{region} {district} {local} 중2 수학학원 선택을 위해 {center} 공개 정보와 "
        f"중2 내신·오답 관리 기준, 수업 가능 학교와 상담 체크 항목을 정리했습니다."
    )
    page_url = canonical(PARENT, CATEGORY, local)
    body_name = "seoul.jpg" if region == "서울" else "local.jpg"
    body_url = f"{DOMAIN}/assets/centers/common/{body_name}"
    map_url = f"{DOMAIN}/assets/maps/{quote(map_name)}"

    parent_link = parent_relative_url(row)
    middle1_link = parent_relative_url(row, "중1수학학원")
    related = [
        (CATEGORY_LABEL, canonical(PARENT, CATEGORY)),
        (f"{local} 중2 영어학원", canonical(PARENT, "중2영어학원", local)),
        (f"{local}학원", parent_canonical_url(row)),
        (f"{local} 중1 수학학원", parent_canonical_url(row, "중1수학학원")),
    ]
    related.extend(
        (f"{peer} 중2 수학학원", canonical(PARENT, CATEGORY, peer))
        for peer in peer_locals
    )
    faqs = build_faqs(row, profile)
    notes = build_parent_notes(row, profile)
    checklist = build_checklist(row, profile)
    schema = page_json_ld(row, description, page_url, rep_image, body_url, map_url, faqs, related)

    school_html = (
        "".join(f"<span>{esc(school)}</span>" for school in schools)
        if schools
        else "<p class=\"subject-empty-note\">공개 자료에 중학교명이 별도로 기재되어 있지 않아 상담 시 현재 학교와 시험 범위를 확인합니다.</p>"
    )
    grade_html = "".join(f"<span>{esc(grade)}</span>" for grade in grades) if grades else "<span>상담 확인</span>"
    availability_text = (
        f"공개 센터 자료의 수학 가능 학년에 중2가 포함되어 있습니다. 실제 반 편성과 일정은 {center} 상담에서 확인해야 합니다."
        if available
        else f"공개된 {center} 자료에는 수학 가능 학년이 비어 있어 중2 수학 개설 여부를 상담에서 확인해야 합니다."
    )
    faq_html = "".join(
        f"<details><summary>{esc(question)}</summary><p>{esc(answer)}</p></details>"
        for question, answer in faqs
    )
    notes_html = "".join(f"<article class=\"subject-parent-note\"><p>{esc(note)}</p></article>" for note in notes)
    checklist_html = "".join(
        f'<article class="geo-check-card"><b>{position:02d}</b><strong>{esc(label)}</strong><p>{esc(body)}</p></article>'
        for position, (label, body) in enumerate(checklist, 1)
    )
    tuition_html = (
        f'<a class="btn ghost" href="{esc(tuition)}" target="_blank" rel="noopener noreferrer">센터 교습비 자료 확인</a>'
        if tuition
        else '<span class="subject-empty-note">교습비 자료는 상담 시 확인해 주세요.</span>'
    )
    tuition_status = "공개 링크 확인 가능" if tuition else "상담 확인 필요"
    location_html = (
        f'<article data-role="verified-location"><span>확인된 위치 안내</span><strong>{esc(center)}</strong><p>{esc(location_guide)}</p></article>'
        if location_guide
        else ""
    )
    peer_links_html = "".join(
        f'<a class="child-page-button" href="../{quote(peer, safe="")}/index.html">{esc(peer)} 중2 수학학원</a>'
        for peer in peer_locals
    )
    manuscript = contextualize_manuscript(manuscript, row, REPEATED_SIGNATURES)
    enrichment = build_math_enrichment(row, profile)

    return f'''{head_html(title, description, page_url, rep_image, schema, "../../../")}
<body>
{nav_html("../../../", "subject")}
  <main id="main">
    <nav class="breadcrumb-box" aria-label="현재 위치"><a href="../../../index.html">홈</a><span>›</span><a href="../../index.html">과목별학원</a><span>›</span><a href="../index.html">중2 수학학원</a><span>›</span>{esc(title)}</nav>

    <section class="sub-hero shell directory-hero subject-detail-hero">
      <div class="reveal">
        <p class="eyebrow">MIDDLE SCHOOL MATH GUIDE</p>
        <h1>{esc(title)}</h1>
        <p>{esc(description)}</p>
        <div class="hero-actions"><a class="btn primary" href="{FORM_URL}" target="_blank" rel="noopener noreferrer">상담 준비하기</a><a class="btn ghost" href="../index.html">다른 지역 찾기</a></div>
      </div>
      <div class="stat-console reveal"><div class="stat-pill"><strong>{esc(local)}</strong><span>{esc(region)} {esc(district)}</span></div><div class="stat-pill"><strong>{'확인' if not available else '중2'}</strong><span>{'개설 여부 상담 필요' if not available else '수학 가능 학년 기재'}</span></div></div>
    </section>

    <section class="shell csv-body-stack csv-top-media local-media-section subject-media" aria-label="{esc(title)} 이미지 안내">
      <img data-role="representative-image" style="display:none;" src="{esc(rep_image)}" alt="{esc(title)} {SITE_NAME} 대표">
      <figure class="csv-media-card"><img src="../../../assets/centers/common/{body_name}" alt="{esc(title)} 본문 {SITE_NAME}" loading="eager" decoding="async"></figure>
      <figure class="csv-media-card"><img src="../../../assets/maps/{quote(map_name)}" alt="{esc(title)} 지도 {SITE_NAME}" loading="lazy" decoding="async"></figure>
    </section>

    <section class="shell geo-summary-panel subject-answer-panel reveal" aria-labelledby="answer-title">
      <p class="eyebrow">핵심 답변</p>
      <h2 id="answer-title">{esc(title)}, 무엇부터 확인해야 할까요?</h2>
      <p>{esc(local)}에서 중2 수학학원을 비교할 때는 현재 진도만 보지 말고 이전 개념 공백, 풀이 과정, 오답 재확인 방식과 시험 전 계획을 함께 살펴야 합니다. {esc(profile['student'])}이라면 특히 관리 기록이 다음 수업에 어떻게 반영되는지 확인하는 것이 좋습니다.</p>
      <div class="geo-fact-grid">
        <article class="geo-fact-card"><span>현재 학생 상황</span><strong>{esc(profile['student'])}</strong></article>
        <article class="geo-fact-card"><span>우선 확인</span><strong>{esc(profile['priority'])}</strong></article>
        <article class="geo-fact-card"><span>상담 질문</span><strong>{esc(profile['check'])}</strong></article>
      </div>
    </section>

    {manuscript}
    {enrichment}

    <section class="shell academy-section subject-local-facts reveal" aria-labelledby="local-facts-title">
      <div class="section-heading"><p class="eyebrow">VERIFIED LOCAL FACTS</p><h2 id="local-facts-title">{esc(local)} 센터 정보와 수업 확인 항목</h2><p>공개된 센터 자료만 사용했으며, 기재되지 않은 학교나 운영 조건은 임의로 만들지 않았습니다.</p></div>
      <div class="subject-fact-grid">
        <article><span>센터</span><strong>{esc(center)}</strong><p>{esc(address) if address else '주소는 상담 시 확인해 주세요.'}</p></article>
        <article><span>중2 수학 가능 학년</span><strong>{'자료에 기재됨' if available else '상담 확인 필요'}</strong><p>{esc(availability_text)}</p></article>
        <article><span>교육지원청 등록 정보</span><strong>{esc(registration_name) if registration_name else '상담 확인'}</strong><p>{esc(registration_number) if registration_number else '공개 자료에 등록번호가 별도로 기재되어 있지 않습니다.'}</p></article>
        <article><span>교습비 확인</span><strong>{esc(tuition_status)}</strong><p>주당 횟수·시간표·보강 기준과 함께 비교합니다.</p></article>
        {location_html}
      </div>
      <div class="subject-school-panel"><h3>공개 자료의 중학교 안내 · {len(schools)}개</h3><div class="subject-school-tags">{school_html}</div></div>
      <div class="subject-grade-panel"><h3>공개 자료의 수학 가능 학년</h3><div class="subject-school-tags">{grade_html}</div>{tuition_html}</div>
    </section>

    <section class="shell geo-checklist-panel reveal" aria-labelledby="checklist-title">
      <p class="eyebrow">상담 전 체크리스트</p>
      <h2 id="checklist-title">{esc(title)} 비교 전에 적어둘 내용</h2>
      <div class="geo-checklist-grid">{checklist_html}</div>
    </section>

    <section class="shell academy-section local-proof-section" aria-labelledby="faq-title">
      <div class="section-heading"><p class="eyebrow">FAQ & PARENT VIEW</p><h2 id="faq-title">{esc(title)} 자주 묻는 질문과 학부모 상담 관점</h2><p>질문과 답변은 같은 내용으로 JSON-LD에도 반영했습니다. 아래 내용은 실제 후기를 가장한 문장이 아니라 학부모가 상담에서 확인할 비교 기준입니다.</p></div>
      <div class="local-proof-layout">
        <section class="local-faq-card" aria-label="{esc(title)} 자주 묻는 질문"><div class="faq-list">{faq_html}</div></section>
        <aside class="local-review-card" aria-label="{esc(title)} 학부모 상담 관점"><div class="review-list">{notes_html}</div></aside>
      </div>
    </section>

    <section class="shell local-page-nav reveal" aria-labelledby="related-title">
      <div class="section-heading"><p class="eyebrow">RELATED GUIDES</p><h2 id="related-title">{esc(local)} 중2 과목과 다른 지역 비교</h2><p>같은 동네 영어 안내와 시군구·광역권을 우선한 수학 페이지를 상호 연결했습니다.</p></div>
      <div class="child-button-grid">
        <a class="child-page-button" href="../index.html">중2 수학학원 지역 목록</a>
        <a class="child-page-button" href="../../중2영어학원/{quote(local, safe='')}/index.html">{esc(local)} 중2 영어학원</a>
        <a class="child-page-button" href="{parent_link}">{esc(local)}학원</a>
        <a class="child-page-button" href="{middle1_link}">{esc(local)} 중1 수학학원</a>
        {peer_links_html}
      </div>
    </section>
  </main>
{footer_html("../../../")}
</body>
</html>
'''


def collection_schema(title: str, description: str, url: str, breadcrumbs: list[tuple[str, str]], items: list[tuple[str, str]], faqs: list[tuple[str, str]] | None = None) -> dict:
    graph: list[dict] = [
        {"@type": "WebSite", "@id": f"{DOMAIN}/#website", "url": f"{DOMAIN}/", "name": SITE_NAME, "inLanguage": "ko-KR"},
        {
            "@type": "CollectionPage",
            "@id": f"{url}#webpage",
            "url": url,
            "name": title,
            "description": description,
            "inLanguage": "ko-KR",
            "isPartOf": {"@id": f"{DOMAIN}/#website"},
            "breadcrumb": {"@id": f"{url}#breadcrumb"},
            "mainEntity": {"@id": f"{url}#itemlist"},
        },
        {
            "@type": "BreadcrumbList",
            "@id": f"{url}#breadcrumb",
            "itemListElement": [
                {"@type": "ListItem", "position": position, "name": name, "item": item_url}
                for position, (name, item_url) in enumerate(breadcrumbs, 1)
            ],
        },
        {
            "@type": "ItemList",
            "@id": f"{url}#itemlist",
            "name": title,
            "numberOfItems": len(items),
            "itemListElement": [
                {"@type": "ListItem", "position": position, "name": name, "url": item_url}
                for position, (name, item_url) in enumerate(items, 1)
            ],
        },
    ]
    if faqs:
        graph.append(
            {
                "@type": "FAQPage",
                "@id": f"{url}#faq",
                "mainEntity": [
                    {"@type": "Question", "name": q, "acceptedAnswer": {"@type": "Answer", "text": a}}
                    for q, a in faqs
                ],
            }
        )
    return {"@context": "https://schema.org", "@graph": graph}


def category_page() -> str:
    title = "과목별학원"
    description = "학년과 과목을 먼저 선택한 뒤 지역별 학습 안내를 찾을 수 있도록 정리한 전문수업.com 과목별 학원 허브입니다."
    url = canonical(PARENT)
    items = [
        (CATEGORY_LABEL, canonical(PARENT, CATEGORY)),
        ("중2 영어학원", canonical(PARENT, "중2영어학원")),
        ("중3 수학학원", canonical(PARENT, "중3수학학원")),
        ("중3 영어학원", canonical(PARENT, "중3영어학원")),
    ]
    schema = collection_schema(title, description, url, [("홈", f"{DOMAIN}/"), (title, url)], items)
    return f'''{head_html(title, description, url, f"{DOMAIN}/assets/generated/academy-hero-v2.webp", schema, "../", "website")}
<body>
{nav_html("../", "subject")}
  <main id="main">
    <nav class="breadcrumb-box" aria-label="현재 위치"><a href="../index.html">홈</a><span>›</span>과목별학원</nav>
    <section class="sub-hero shell directory-hero">
      <div class="reveal"><p class="eyebrow">SUBJECT ACADEMY DIRECTORY</p><h1>과목별학원</h1><p>{esc(description)}</p></div>
      <div class="stat-console reveal"><div class="stat-pill"><strong>학년·과목</strong><span>검색 의도에 맞춘 안내</span></div><div class="stat-pill"><strong>371</strong><span>지역별 상세 연결</span></div></div>
    </section>
    <section class="shell academy-section subject-category-section">
      <div class="section-heading"><p class="eyebrow">CURRENT GUIDE</p><h2>현재 확인할 수 있는 학년·과목</h2><p>실제로 생성된 허브만 표시하며, 존재하지 않는 카테고리는 노출하지 않습니다.</p></div>
      <div class="subject-category-grid">
        <a class="subject-category-card" href="중2수학학원/index.html"><span>중학교 2학년</span><strong>중2 수학학원</strong><p>내신 준비, 개념 공백, 풀이 과정과 오답 재학습 기준을 지역별로 확인합니다.</p><b>371개 지역 보기 →</b></a>
        <a class="subject-category-card" href="중2영어학원/index.html"><span>중학교 2학년</span><strong>중2 영어학원</strong><p>교과서·문법·독해·어휘와 서술형 준비 기준을 지역별로 확인합니다.</p><b>371개 지역 보기 →</b></a>
        <a class="subject-category-card" href="중3수학학원/index.html"><span>중학교 3학년</span><strong>중3 수학학원</strong><p>내신·서술형·누적 공백과 고교 진입 전 수학 준비 기준을 지역별로 확인합니다.</p><b>371개 지역 보기 →</b></a>
        <a class="subject-category-card" href="중3영어학원/index.html"><span>중학교 3학년</span><strong>중3 영어학원</strong><p>교과서 내신, 누적 어휘·문법·독해와 고교 진입 전 영어 준비 기준을 지역별로 확인합니다.</p><b>371개 지역 보기 →</b></a>
      </div>
    </section>
  </main>
{footer_html("../")}
</body>
</html>
'''


def hub_page(rows: list[dict[str, str]]) -> str:
    title = "중2 수학학원 지역별 안내"
    description = "371개 지역의 중2 수학학원 선택 기준과 공개 센터 정보, 학교·가능 학년·교습비 확인 항목을 지역별로 정리했습니다."
    url = canonical(PARENT, CATEGORY)
    items = [
        (f"{row_value(row, '근처 수업가능 동네')} 중2 수학학원", canonical(PARENT, CATEGORY, row_value(row, "근처 수업가능 동네")))
        for row in rows
    ]
    faqs = [
        ("지역별 중2 수학학원 페이지는 어떤 기준으로 구성했나요?", "센터정보 정리 자료의 지역, 센터명, 주소, 실제 기재 학교, 가능 학년과 교습비 링크를 기준으로 구성했습니다."),
        ("수학 가능 학년이 비어 있는 지역도 있나요?", "네. 공개 자료에 수학 가능 학년이 없는 지역은 개설을 임의로 단정하지 않고 상담 확인 필요로 표시했습니다."),
        ("동네 이름으로 바로 찾을 수 있나요?", "검색창에 동네명, 시군구 또는 센터명을 입력하면 해당 지역 버튼만 확인할 수 있습니다."),
    ]
    schema = collection_schema(
        title,
        description,
        url,
        [("홈", f"{DOMAIN}/"), (PARENT, canonical(PARENT)), (CATEGORY_LABEL, url)],
        items,
        faqs,
    )
    grouped: dict[str, dict[str, list[dict[str, str]]]] = defaultdict(lambda: defaultdict(list))
    for row in rows:
        grouped[row_value(row, "지역")][row_value(row, "시or구")].append(row)

    region_sections: list[str] = []
    for region_index, (region, districts) in enumerate(grouped.items()):
        count = sum(len(values) for values in districts.values())
        district_sections: list[str] = []
        for district, district_rows in districts.items():
            cards = []
            for row in district_rows:
                local = row_value(row, "근처 수업가능 동네")
                center = row_value(row, "센터명")
                schools = row_value(row, "타깃학교\n(중)")
                search = " ".join([region, district, local, center, schools])
                cards.append(
                    f'<a class="subject-town-card" data-subject-town data-search="{esc(search)}" href="{quote(local, safe="")}/index.html"><strong>{esc(local)}</strong><span>{esc(district)} · 중2 수학</span></a>'
                )
            district_sections.append(
                f'<section class="subject-district-group" data-subject-district><h3>{esc(district)} <small>{len(district_rows)}개 지역</small></h3><div class="subject-town-grid">{"".join(cards)}</div></section>'
            )
        region_sections.append(
            f'<details class="subject-region-group" data-subject-region{" open" if region_index == 0 else ""}><summary><span>{esc(region)}</span><b>{count}개 지역</b></summary><div class="subject-region-body">{"".join(district_sections)}</div></details>'
        )
    faq_html = "".join(f"<details><summary>{esc(q)}</summary><p>{esc(a)}</p></details>" for q, a in faqs)
    return f'''{head_html(title, description, url, f"{DOMAIN}/assets/generated/academy-hero-v2.webp", schema, "../../", "website")}
<body>
{nav_html("../../", "subject")}
  <main id="main">
    <nav class="breadcrumb-box" aria-label="현재 위치"><a href="../../index.html">홈</a><span>›</span><a href="../index.html">과목별학원</a><span>›</span>중2 수학학원</nav>
    <section class="sub-hero shell directory-hero">
      <div class="reveal"><p class="eyebrow">MIDDLE SCHOOL MATH DIRECTORY</p><h1>중2 수학학원</h1><p>{esc(description)}</p></div>
      <div class="stat-console reveal"><div class="stat-pill"><strong>371</strong><span>동네별 학습 안내</span></div><div class="stat-pill"><strong>13</strong><span>광역 지역 구분</span></div></div>
    </section>
    <section class="shell academy-section subject-directory" data-subject-directory>
      <div class="section-heading"><p class="eyebrow">LOCAL SEARCH</p><h2>동네 이름으로 중2 수학학원 찾기</h2><p>광역·시군구별로 접어 정리했습니다. 동네명이나 센터명을 검색하면 해당 결과만 남습니다.</p></div>
      <div class="subject-search-box"><label for="subject-town-search">동네·시군구·센터 검색</label><input id="subject-town-search" type="search" placeholder="예: 명일동, 강동구, 명일점" autocomplete="off" data-subject-search><p aria-live="polite" data-subject-search-status>전체 371개 지역</p></div>
      <div class="subject-region-list">{"".join(region_sections)}</div>
    </section>
    <section class="shell academy-section subject-hub-faq"><div class="section-heading"><p class="eyebrow">DIRECTORY FAQ</p><h2>지역 페이지 이용 전에 확인하세요</h2></div><div class="faq-list">{faq_html}</div></section>
  </main>
{footer_html("../../")}
</body>
</html>
'''


def update_sitemap(urls: list[tuple[str, str, str]]) -> None:
    path = SITE / "sitemap.xml"
    text = path.read_text(encoding="utf-8")
    existing = set(re.findall(r"<loc>(.*?)</loc>", text))
    additions: list[str] = []
    for url, changefreq, priority in urls:
        if url in existing:
            continue
        additions.append(
            "  <url>\n"
            f"    <loc>{url}</loc>\n"
            f"    <lastmod>{PUBLISH_DATE}</lastmod>\n"
            f"    <changefreq>{changefreq}</changefreq>\n"
            f"    <priority>{priority}</priority>\n"
            "  </url>\n"
        )
    if additions:
        text = text.replace("</urlset>", "".join(additions) + "</urlset>")
        path.write_text(text, encoding="utf-8", newline="\n")


def main() -> None:
    global REPEATED_SIGNATURES
    rows = read_csv(CENTER_CSV)
    manuscripts = load_manuscripts()
    REPEATED_SIGNATURES = repeated_paragraphs(manuscripts, rows)
    if len(rows) != 371 or len(manuscripts) != 371:
        raise ValueError(f"371개 자료 불일치: center={len(rows)}, manuscript={len(manuscripts)}")
    if len({row_value(row, '근처 수업가능 동네') for row in rows}) != 371:
        raise ValueError("동네명이 중복되었습니다.")

    image_rows = load_image_rows()
    fallback_representatives = representative_urls()
    if not fallback_representatives:
        raise ValueError("대표 이미지 URL을 찾지 못했습니다.")

    target = SITE / PARENT / CATEGORY
    target.mkdir(parents=True, exist_ok=True)
    (SITE / PARENT).mkdir(parents=True, exist_ok=True)
    (SITE / PARENT / "index.html").write_text(category_page(), encoding="utf-8", newline="\n")
    (target / "index.html").write_text(hub_page(rows), encoding="utf-8", newline="\n")
    peer_network = build_subject_peer_network(rows)

    sitemap_urls = [
        (canonical(PARENT), "weekly", "0.9"),
        (canonical(PARENT, CATEGORY), "weekly", "0.9"),
    ]
    for index, (row, manuscript) in enumerate(zip(rows, manuscripts)):
        local = row_value(row, "근처 수업가능 동네")
        image_row = image_rows.get(local, {})
        map_name = map_filename(row, image_row)
        rep_image = representative_for(row, fallback_representatives)
        folder = target / local
        folder.mkdir(parents=True, exist_ok=True)
        page = detail_html(
            row,
            manuscript,
            image_row,
            rep_image,
            map_name,
            index,
            rows,
            peer_network[local],
        )
        (folder / "index.html").write_text(clean_markup(page), encoding="utf-8", newline="\n")
        sitemap_urls.append((canonical(PARENT, CATEGORY, local), "monthly", "0.8"))

    update_sitemap(sitemap_urls)
    print(f"generated_details={len(rows)}")
    print(f"generated_total={len(rows) + 2}")
    print(f"target={target}")


REPEATED_SIGNATURES: set[str] = set()


if __name__ == "__main__":
    main()
